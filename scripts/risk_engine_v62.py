# -*- coding: utf-8 -*-
"""
IT Risk Engine v5.1
Tum ciktilar: Excel (11 sheet) + PPTX Alert (7 slayt) + Outlook Mail (Display modu)
SORUN GIDERME: Script kapanirsa logs/error.log dosyasini kontrol edin.
"""

import os, sys, json, traceback, subprocess
from datetime import datetime

# ── Hata ayni anda log a yazilir, CMD kapanmaz ──────────────────
LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "engine.log")

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as lf:
        lf.write(line + "\n")

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    import win32com.client as win32
    WIN32_OK = True
except ImportError:
    win32 = None
    WIN32_OK = False

# ── KONFIGÜRASYON ────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
BASE_DIR    = os.path.dirname(SCRIPT_DIR)
INPUT_FILE  = os.path.join(BASE_DIR, "data", "raw", "lansweeper_risk.xlsx")
REPORTS_DIR = os.path.join(BASE_DIR, "output", "reports")
ALERTS_DIR  = os.path.join(BASE_DIR, "output", "alerts")
TEMP_DIR    = os.path.join(BASE_DIR, "output", "_tmp")
for d in [REPORTS_DIR, ALERTS_DIR, TEMP_DIR, LOG_DIR]:
    os.makedirs(d, exist_ok=True)

TARIH   = datetime.now().strftime("%Y%m%d")
TARIH_H = datetime.now().strftime("%d.%m.%Y")
OUTPUT_EXCEL = os.path.join(REPORTS_DIR, f"IT_RISK_REPORT_{TARIH}.xlsx")
OUTPUT_PPTX  = os.path.join(ALERTS_DIR,  f"IT_RISK_ALERT_{TARIH}.pptx")

PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
os.makedirs(PROCESSED_DIR, exist_ok=True)
# Mail alıcıları ortam değişkenlerinden okunur (virgülle ayrılmış liste).
# Örnek .env:
#   RISK_MAIL_TO="it-team@example.com,manager@example.com"
#   RISK_MAIL_CC="director@example.com"
MAIL_TO = [x.strip() for x in os.getenv("RISK_MAIL_TO", "").split(",") if x.strip()]
MAIL_CC = [x.strip() for x in os.getenv("RISK_MAIL_CC", "").split(",") if x.strip()]

# Renk sabitleri
CR="E74C3C"; CO="F39C12"; CG="27AE60"; CB="2980B9"
CD="2C3E50"; CW="FFFFFF"; CP="8E44AD"; BG="#12181F"

try:
    # ═══════════════════════════════════════════════════════════
    # 1. VERİ YÜKLEME
    # ═══════════════════════════════════════════════════════════
    log("="*60)
    log("IT Risk Engine v5.1 Baslatiyor...")
    log("="*60)

    if not os.path.exists(INPUT_FILE):
        log(f"HATA: {INPUT_FILE} bulunamadi!"); sys.exit(1)

    df = pd.read_excel(INPUT_FILE)
    df.columns = df.columns.str.strip()
    log(f"Sutunlar: {list(df.columns)}")

    # Sayisal sutunlari temizle
    for col in df.columns:
        c = col.strip()
        if c in ["% Bos","Offline Gun","Yamasiz Gun","Risk Skoru",
                 "_RawDiskError","_RawAdminCount","_RawUpdateStop",
                 "% Boş","Offline Gün","Yamasız Gün"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ── AssetID sütun normalizasyonu ──────────────────────────────────────────
    # Lansweeper'dan gelen dosyada sütun adı "AssetID1", "AssetID", "Asset ID"
    # gibi farklı varyantlarda olabilir. Tümünü "AssetID" standart adına çevir.
    import re as _re
    def _norm_col(s):
        return _re.sub(r"[^a-z0-9]", "", s.strip().lower())
    _AID_VARIANTS = {"assetid", "assetid1", "assetid2", "asset id", "asset_id", "id"}
    _aid_src = next(
        (c for c in df.columns if _norm_col(c) in _AID_VARIANTS
         and _norm_col(c) != "id"),   # "id" tek başına çok genel, son çare
        next((c for c in df.columns if _norm_col(c) == "id"), None)
    )
    if _aid_src and _aid_src != "AssetID":
        df["AssetID"] = df[_aid_src].copy()
        log(f"AssetID normalizasyonu: '{_aid_src}' → 'AssetID'")
    elif "AssetID" not in df.columns:
        df["AssetID"] = ""
        log("UYARI: AssetID sütunu bulunamadı — Lansweeper linkleri çalışmayabilir!")

    # AssetID'yi integer string olarak temizle (12345.0 → 12345)
    def _clean_aid(val):
        try:
            v = str(val).strip()
            if v in ("", "nan", "None", "NaN", "0", "0.0"): return ""
            return str(int(float(v)))
        except Exception:
            return str(val).strip() if str(val).strip() not in ("nan","None") else ""
    df["AssetID"] = df["AssetID"].apply(_clean_aid)
    log(f"AssetID örnek değerler: {df['AssetID'].dropna().head(5).tolist()}")

    df["Kullanıcı"]    = df["Kullanıcı"].fillna("Sahipsiz") if "Kullanıcı" in df.columns else "Sahipsiz"
    df["Risk Analizi"] = df["Risk Analizi"].fillna("Risk Tespit Edilmedi")

    # ── KRİTİK DÜZELTME: Seviye sütununa güvenme, Risk Skoru'ndan yeniden hesapla ──
    df["Risk Skoru"] = pd.to_numeric(df["Risk Skoru"], errors="coerce").fillna(0)

    # ═══════════════════════════════════════════════════════════
    # ASSET CRİTİCALİTY MODELİ
    # Lansweeper skoruna dokunmuyoruz. Üstüne çarpan uyguluyoruz.
    # Final_Risk_Skoru = min(100, Lansweeper_Skor × Çarpan + Bonus)
    # ═══════════════════════════════════════════════════════════
    # Cihaz tipi tespiti: AssetName veya Sistem sütunundaki anahtar kelimelere göre
    CRITICALITY_RULES = [
        (r"(?i)\bDC\b|\bDC\d+\b|DOMAIN.CTRL|-DC-",  r"(?i)Server",  "Domain Controller",    1.6, 25),
        (r"(?i)EXCH|MAIL.SRV|-EX\d",                 r"(?i)Server",  "Mail Server",           1.5, 20),
        (r"(?i)SQL|DB.SRV|-DB\d",                     r"(?i)Server",  "Veritabanı Sunucusu",  1.5, 20),
        (r"(?i)SRV|SVR|SERVER",                       r"",            "Sunucu",                1.4, 15),
        (r"",                                          r"(?i)Server",  "Sunucu (OS)",           1.35, 10),
        (r"(?i)LAP|NTB|NOTEBOOK",                     r"",            "Laptop",                1.0,  0),
        (r"",                                          r"",            "Workstation",           1.0,  0),
    ]

    import re as _re

    def _get_criticality(asset_name, sistem):
        an      = str(asset_name) if pd.notna(asset_name) else ""
        sys_val = str(sistem)     if pd.notna(sistem)     else ""
        for pat_name, pat_sys, label, mult, bonus in CRITICALITY_RULES:
            m_name = bool(pat_name and _re.search(pat_name, an))
            m_sys  = bool(pat_sys  and _re.search(pat_sys,  sys_val))
            if pat_name and pat_sys:
                if m_name and m_sys: return label, mult, bonus
            elif pat_name:
                if m_name: return label, mult, bonus
            elif pat_sys:
                if m_sys:  return label, mult, bonus
            else:
                return label, mult, bonus
        return "Workstation", 1.0, 0

    crit_data = df.apply(
        lambda r: _get_criticality(
            r.get("AssetName", r.get("AssetID", "")),
            r.get("Sistem", "")
        ), axis=1
    )
    df["Cihaz_Tipi"]     = crit_data.apply(lambda x: x[0])
    df["Crit_Multiplier"]= crit_data.apply(lambda x: x[1])
    df["Crit_Bonus"]     = crit_data.apply(lambda x: x[2])

    # Final skor: çarpan uygula, 100 ile sınırla
    df["Final_Risk_Skoru"] = (df["Risk Skoru"] * df["Crit_Multiplier"] + df["Crit_Bonus"]).clip(upper=100).round(0).astype(int)

    # Seviye artık Final skor üzerinden belirlenir
    df["Seviye"] = df["Final_Risk_Skoru"].apply(
        lambda x: "YUKSEK" if x >= 50 else ("ORTA" if x >= 25 else "DUSUK")
    )

    # Criticality istatistikleri
    crit_counts = df["Cihaz_Tipi"].value_counts().to_dict()
    n_servers   = int(df["Cihaz_Tipi"].str.contains("Sunucu|Controller|Server|Mail|Veri", na=False).sum())
    n_promoted  = int((df["Final_Risk_Skoru"] > df["Risk Skoru"]).sum())  # kaç cihaz yukarı çıktı

    log(f"Criticality: {crit_counts}")
    log(f"  → {n_servers} sunucu/kritik varlık tespit edildi, {n_promoted} cihaz skoru yukarı çıktı.")

    n_total  = len(df)
    n_high   = int((df["Seviye"] == "YUKSEK").sum())
    n_medium = int((df["Seviye"] == "ORTA").sum())
    n_low    = int((df["Seviye"] == "DUSUK").sum())

    log(f"Veri yuklendi: {n_total} cihaz | Yuksek:{n_high} Orta:{n_medium} Dusuk:{n_low} (Criticality uygulandı)")

    # ═══════════════════════════════════════════════════════════
    # 2. ANALİZ
    # ═══════════════════════════════════════════════════════════
    threat_map = {
        "DLP Yüklü Değil":          "DLP / Uç Nokta Güvenliği Eksik",
        "Onaysız Yönetici Yetkisi": "Yetkisiz Admin Hesabı",
        "Şüpheli Yazılım":          "Shadow IT / Yasaklı Yazılım",
        "Desteklenmeyen OS":        "EoL İşletim Sistemi",
        "Antivirüs (SEP) Eksik":    "Antivirüs Koruması Yok",
        "Güvenlik Yamaları Eksik":  "Patch / Yama Eksikliği",
        "Update Servisi Kapalı":    "Windows Update Servisi Kapalı",
        "Uzun Süredir Offline":     "Zombi / Kayıp Cihaz",
        "Riskli Paylaşım":          "Açık Klasör Paylaşımı",
        "Kritik Disk":              "Kritik Disk Doluluk / Arızası",
    }
    threat_stats = {}
    for kw, lbl in threat_map.items():
        c = int(df["Risk Analizi"].astype(str).str.contains(kw, na=False, regex=False).sum())
        if c > 0:
            threat_stats[lbl] = c
    threat_df = (pd.DataFrame(list(threat_stats.items()), columns=["Tehdit","Sayi"])
                   .sort_values("Sayi", ascending=False).reset_index(drop=True))

    # Kullanıcı risk puanı — Final skor kullan
    user_risk = df.groupby("Kullanıcı").agg(
        Cihaz_Sayisi    = ("Final_Risk_Skoru", "count"),
        Ort_Risk_Skoru  = ("Final_Risk_Skoru", "mean"),
        En_Yuksek_Skor  = ("Final_Risk_Skoru", "max"),
        Yuksek_Riskli   = ("Seviye",           lambda x: (x=="YUKSEK").sum()),
        Max_Yamasiz_Gun = ("Yamasız Gün",      "max"),
        Toplam_Admin    = ("_RawAdminCount",   "sum"),
        Disk_Hata_Sayi  = ("_RawDiskError",    "sum"),
        WU_Kapali_Sayi  = ("_RawUpdateStop",   "sum"),
    ).reset_index()
    user_risk["Davraniş_Risk_Puani"] = (
        user_risk["Ort_Risk_Skoru"]
      + user_risk["Toplam_Admin"]   * 10
      + user_risk["Max_Yamasiz_Gun"].apply(lambda x: 20 if x > 60 else 0)
      + user_risk["Disk_Hata_Sayi"] * 5
      + user_risk["WU_Kapali_Sayi"] * 5
    ).round(1)
    user_risk = user_risk.sort_values("Davraniş_Risk_Puani", ascending=False).head(15)

    eol_count    = int(df["Sistem"].str.contains("Win 7|2008|8.1|XP|2012", na=False).sum())
    patch_crit   = int((df["Yamasız Gün"] > 60).sum())
    off_crit     = int((df["Offline Gün"] > 60).sum())
    disk_crit    = int((df["% Boş"] < 10).sum())
    admin_crit   = int(df["_RawAdminCount"].gt(0).sum())
    shadow_crit  = int(df["Tespit Edilen Şüpheli Yazılımlar"].notna().sum()) \
                   if "Tespit Edilen Şüpheli Yazılımlar" in df.columns else 0
    share_crit   = int(df["Riskli Paylaşılan Klasörler"].notna().sum()) \
                   if "Riskli Paylaşılan Klasörler" in df.columns else 0
    avg_score    = round(float(df["Final_Risk_Skoru"].mean()), 1)
    max_score    = int(df["Final_Risk_Skoru"].max())

    log(f"Analiz: EoL={eol_count} Patch={patch_crit} Offline={off_crit} Disk={disk_crit} Admin={admin_crit} Shadow={shadow_crit}")

    # ═══════════════════════════════════════════════════════════
    # CVE VERİSİ ENTEGRASYONU
    # cve_scanner.py önce çalıştırılmışsa, en güncel JSON dosyasını yükle
    # ve risk motoruna CVE bonusu uygula
    # ═══════════════════════════════════════════════════════════
    CVE_DATA      = {}    # {"yazilim_adi": {"max_cvss":x, "bonus":y, ...}}
    CVE_META      = {}    # tarama istatistikleri
    cve_file_used = None

    # En güncel cve_summary_*.json dosyasını bul
    import glob as _glob
    cve_jsons = sorted(
        _glob.glob(os.path.join(PROCESSED_DIR, "cve_summary_*.json")),
        reverse=True  # en yeni başta
    )
    if cve_jsons:
        cve_file_used = cve_jsons[0]
        try:
            with open(cve_file_used, encoding="utf-8") as f:
                cve_raw = json.load(f)
            CVE_DATA = cve_raw.get("sw_risk", {})
            CVE_META = {k: v for k, v in cve_raw.items() if k != "sw_risk"}
            log(f"CVE verisi yüklendi: {cve_file_used}")
            log(f"  → {CVE_META.get('vuln_yazilim',0)} açık yazılım, "
                f"{CVE_META.get('toplam_cve',0)} CVE, "
                f"{CVE_META.get('kritik',0)} kritik")

            # Şüpheli yazılım sütunundan CVE bonus uygula
            # CVE_DATA anahtarlarını lowercase index'e çevir (hızlı arama için)
            cve_index = {k.lower().strip(): v for k, v in CVE_DATA.items()}

            def _best_cve_match(sw_name: str) -> int:
                """Yazılım adı için en yüksek CVE bonusunu döndürür."""
                key = sw_name.lower().strip()
                if not key or len(key) < 3:
                    return 0
                # 1. Tam eşleşme
                if key in cve_index:
                    return cve_index[key].get("bonus", 0)
                # 2. Kısmi eşleşme (key, CVE adının içinde veya tam tersi)
                for cve_key, cve_val in cve_index.items():
                    if key in cve_key or cve_key in key:
                        return cve_val.get("bonus", 0)
                return 0

            def _cve_bonus_from_cell(cell) -> int:
                """Şüpheli yazılım hücresi için bonus (pipe ile ayrılmış)."""
                if not isinstance(cell, str) or not cell.strip():
                    return 0
                return max(
                    (_best_cve_match(p.strip()) for p in cell.split("|")),
                    default=0
                )

            col_sw = "Tespit Edilen Şüpheli Yazılımlar"

            # Yöntem 1: Şüpheli yazılım sütunu
            if col_sw in df.columns:
                df["CVE_Bonus"] = df[col_sw].apply(_cve_bonus_from_cell)
            else:
                df["CVE_Bonus"] = 0

            # Yöntem 2: software_inventory.xlsx varsa cihaz-yazılım eşleştirmesi
            sw_inv_path = os.path.join(BASE_DIR, "data", "raw", "software_inventory.xlsx")
            if os.path.exists(sw_inv_path):
                try:
                    df_sw_inv = pd.read_excel(sw_inv_path)
                    # Her cihaz için en yüksek CVE bonusunu bul
                    sw_bonus_map = {}  # AssetName → max bonus
                    for _, row in df_sw_inv.iterrows():
                        asset = str(row.get("AssetName", "")).strip()
                        sw    = str(row.get("SoftwareName", "")).strip()
                        if not asset or not sw:
                            continue
                        bonus = _best_cve_match(sw)
                        if bonus > 0:
                            sw_bonus_map[asset] = max(sw_bonus_map.get(asset, 0), bonus)
                    # DataFrame'e uygula
                    def _inv_bonus(asset_name):
                        return sw_bonus_map.get(str(asset_name).strip(), 0)
                    inv_bonus_series = df["AssetName"].apply(_inv_bonus) if "AssetName" in df.columns else pd.Series(0, index=df.index)
                    # İkisinden büyüğünü al
                    df["CVE_Bonus"] = df["CVE_Bonus"].combine(inv_bonus_series, max)
                    n_inv_matched = int((inv_bonus_series > 0).sum())
                    log(f"  → software_inventory.xlsx: {n_inv_matched} cihaz CVE eşleşmesi bulundu")
                except Exception:
                    log(f"  software_inventory eşleştirme hatası (önemsiz): {traceback.format_exc()[:200]}")

            n_cve_affected = int((df["CVE_Bonus"] > 0).sum())

            # Debug: eşleşen/eşleşmeyen örnekleri logla
            if n_cve_affected == 0 and col_sw in df.columns:
                sample_sw = df[col_sw].dropna().head(5).tolist()
                log(f"  [DEBUG] Şüpheli yazılım örnekleri: {sample_sw}")
                log(f"  [DEBUG] CVE_DATA ilk 5 anahtar: {list(CVE_DATA.keys())[:5]}")
            else:
                matched = df[df["CVE_Bonus"]>0][["AssetName","CVE_Bonus"]].head(5).to_dict("records") if "AssetName" in df.columns else []
                log(f"  [DEBUG] Örnek eşleşmeler: {matched}")

            # CVE bonusunu Final_Risk_Skoru'na ekle (100'ü geçmez)
            df["Final_Risk_Skoru"] = (
                df["Final_Risk_Skoru"] + df["CVE_Bonus"]
            ).clip(upper=100).astype(int)

            # Seviyeyi yeniden hesapla
            df["Seviye"] = df["Final_Risk_Skoru"].apply(
                lambda x: "YUKSEK" if x >= 50 else ("ORTA" if x >= 25 else "DUSUK")
            )
            # Sayıları güncelle
            n_high   = int((df["Seviye"] == "YUKSEK").sum())
            n_medium = int((df["Seviye"] == "ORTA").sum())
            n_low    = int((df["Seviye"] == "DUSUK").sum())

            log(f"  → CVE bonusu {n_cve_affected} cihaza uygulandı → "
                f"Yüksek:{n_high} Orta:{n_medium} Düşük:{n_low}")

        except Exception:
            log(f"CVE verisi yüklenemedi (önemsiz):\n{traceback.format_exc()}")
            df["CVE_Bonus"] = 0
    else:
        log("CVE verisi bulunamadı — cve_scanner.py henüz çalıştırılmamış.")
        log("  → Çalıştırmak için: python scripts/cve_scanner.py")
        df["CVE_Bonus"] = 0

    cve_total   = CVE_META.get("toplam_cve", 0)
    cve_kritik  = CVE_META.get("kritik", 0)
    cve_vuln_sw = CVE_META.get("vuln_yazilim", 0)
    # ═══════════════════════════════════════════════════════════
    log("AI Güvenlik Yorumu üretiliyor (Claude API)...")

    CLAUDE_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    # API anahtarını buraya da yazabilirsiniz (önerilmez, env var kullanın):
    # CLAUDE_API_KEY = "sk-ant-..."

    AI_COMMENT_TR  = ""   # Türkçe tam yorum
    AI_BULLETS_TR  = []   # Madde madde öneriler listesi
    AI_SUBJECT_LINE = ""  # Mail konu satırı tamamlayıcısı

    if CLAUDE_API_KEY:
        try:
            import urllib.request as _req

            # En riskli 5 cihazı özetle
            top5 = df.sort_values("Final_Risk_Skoru", ascending=False).head(5)
            top5_txt = "\n".join([
                f"  - {r.get('AssetName','?')} | {r.get('Cihaz_Tipi','?')} | "
                f"Skor:{int(r.get('Final_Risk_Skoru',0))} | {str(r.get('Risk Analizi',''))[:80]}"
                for _, r in top5.iterrows()
            ])

            # En riskli 3 kullanıcı
            top3u = user_risk.head(3)
            top3u_txt = "\n".join([
                f"  - {r['Kullanıcı']} | {int(r['Cihaz_Sayisi'])} cihaz | Puan:{r['Davraniş_Risk_Puani']:.0f}"
                for _, r in top3u.iterrows()
            ])

            threat_txt = "\n".join([
                f"  - {r['Tehdit']}: {r['Sayi']} cihaz"
                for _, r in threat_df.iterrows()
            ])

            prompt = f"""Sen kıdemli bir IT Güvenlik Analisti ve risk raporlama uzmanısın.
Aşağıda bu haftanın otomatik oluşturulan siber güvenlik raporu verileri var.
Bu verileri analiz ederek yöneticilere sunulacak profesyonel bir Türkçe güvenlik yorumu oluştur.

=== HAFTALIK VERİLER ({TARIH_H}) ===
Toplam cihaz: {n_total}
Yüksek riskli: {n_high} (%{round(n_high/n_total*100,1) if n_total else 0})
Orta riskli: {n_medium} (%{round(n_medium/n_total*100,1) if n_total else 0})
Düşük riskli: {n_low} (%{round(n_low/n_total*100,1) if n_total else 0})
Ortalama risk skoru: {avg_score}/100
En yüksek risk skoru: {max_score}/100
Yamasız cihaz (>60 gün): {patch_crit}
Offline cihaz (>60 gün): {off_crit}
Yetkisiz admin: {admin_crit} cihaz
EoL OS (Win7/2008/8.1): {eol_count} cihaz
Shadow IT / yasaklı yazılım: {shadow_crit} cihaz
Kritik disk (<10% boş): {disk_crit} cihaz
Riskli klasör paylaşımı: {share_crit}
Tespit edilen sunucu/kritik varlık: {n_servers}
Criticality nedeniyle skoru yukarı çıkan cihaz: {n_promoted}

CVE bilgileri (NIST NVD taraması):
Taranan yazılım: {CVE_META.get('toplam_tarama', 'Taranmadı')}
Açık bulunan yazılım: {cve_vuln_sw}
Toplam CVE (≥7.0): {cve_total}
Kritik CVE (≥9.0): {cve_kritik}
{threat_txt}

En riskli 5 cihaz:
{top5_txt}

En riskli 3 kullanıcı:
{top3u_txt}

=== GÖREV ===
JSON formatında yanıt ver. Başka hiçbir şey yazma. Kesinlikle JSON dışında metin ekleme.
Format:
{{
  "ozet_paragraf": "3-4 cümlelik yönetici özeti. Durumu bütünsel değerlendir. Geçen haftayla kıyaslama yap (veri olmadığı için 'Bu haftaki tablo gösteriyor ki...' ile başla). Kritik riskleri vurgula. Türkçe, net, profesyonel.",
  "kritik_bulgular": ["Madde 1 (en kritik bulgu, somut rakamlarla)", "Madde 2", "Madde 3"],
  "acil_aksiyonlar": ["Aksiyon 1 (bugün yapılmalı)", "Aksiyon 2 (bu hafta)", "Aksiyon 3"],
  "olumlu_gozlemler": ["Olumlu nokta 1 (varsa)", "Olumlu nokta 2 (varsa)"],
  "risk_trend_yorum": "Risk trendine dair 1-2 cümle yorum. Hangi tehdide odaklanılmalı?",
  "mail_konu_eki": "Mail konusuna eklenecek kısa İngilizce/Türkçe özet (örn: '| 3 kritik sunucu risk altında')"
}}"""

            body = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 1000,
                "messages": [{"role": "user", "content": prompt}]
            }).encode("utf-8")

            req = _req.Request(
                "https://api.anthropic.com/v1/messages",
                data=body,
                headers={
                    "Content-Type":      "application/json",
                    "x-api-key":         CLAUDE_API_KEY,
                    "anthropic-version": "2023-06-01",
                },
                method="POST"
            )
            with _req.urlopen(req, timeout=30) as resp:
                raw = json.loads(resp.read().decode("utf-8"))

            ai_text = raw["content"][0]["text"].strip()
            # JSON fence varsa temizle
            if ai_text.startswith("```"):
                ai_text = ai_text.split("```")[1]
                if ai_text.startswith("json"): ai_text = ai_text[4:]
            ai_json = json.loads(ai_text.strip())

            AI_COMMENT_TR   = ai_json.get("ozet_paragraf", "")
            AI_BULLETS_TR   = (
                [f"🔴 {b}" for b in ai_json.get("kritik_bulgular", [])] +
                [f"⚡ {b}" for b in ai_json.get("acil_aksiyonlar", [])] +
                [f"✅ {b}" for b in ai_json.get("olumlu_gozlemler", [])]
            )
            AI_RISK_TREND   = ai_json.get("risk_trend_yorum", "")
            AI_SUBJECT_LINE = ai_json.get("mail_konu_eki", "")

            log(f"AI Yorum OK: {AI_COMMENT_TR[:80]}...")

        except Exception:
            log(f"AI YORUM HATASI (önemsiz, devam edilecek):\n{traceback.format_exc()}")
            AI_COMMENT_TR = (f"Bu hafta {n_total} cihaz analiz edildi. "
                             f"{n_high} yüksek riskli cihaz tespit edildi "
                             f"(oran: %{round(n_high/n_total*100,1) if n_total else 0}). "
                             f"Ortalama risk skoru {avg_score}/100. "
                             f"Acil öncelik: {admin_crit} yetkisiz admin hesabı ve {patch_crit} yamasız cihaz.")
            AI_BULLETS_TR = [
                f"🔴 {admin_crit} cihazda yetkisiz admin hesabı — Acil AD müdahalesi gerekiyor.",
                f"🔴 {patch_crit} cihaz 60+ gün yamasız — Fidye yazılımı saldırısına açık.",
                f"🔴 {eol_count} cihazda desteksiz OS — Artık güvenlik yaması çıkmıyor.",
                f"⚡ Shadow IT: {shadow_crit} cihazda yasaklı yazılım tespit edildi.",
                f"⚡ {n_servers} kritik sunucu/DC tespit edildi — Skora bonus eklendi.",
            ]
            AI_RISK_TREND   = f"En kritik tehdit: Patch uyumsuzluğu ({patch_crit} cihaz). Sunucu sınıfı cihazlar öncelikli izlenmeli."
            AI_SUBJECT_LINE = ""
    else:
        log("ANTHROPIC_API_KEY bulunamadı — AI yorum atlanıyor. Kural tabanlı yorum üretiliyor.")
        AI_COMMENT_TR = (f"Bu hafta {n_total} cihaz analiz edildi. "
                         f"{n_high} yüksek riskli cihaz tespit edildi "
                         f"(oran: %{round(n_high/n_total*100,1) if n_total else 0}). "
                         f"Ortalama risk skoru {avg_score}/100 olarak ölçüldü. "
                         f"{n_servers} kritik sunucu/DC varlığı tespit edildi ve bu cihazlara ek risk çarpanı uygulandı. "
                         f"Öncelikli eylem: {admin_crit} yetkisiz admin hesabının kaldırılması ve {patch_crit} yamasız cihazın güncellenmesi.")
        AI_BULLETS_TR = [
            f"🔴 {n_high} yüksek riskli cihaz — Acil inceleme listesi raporun 1. sayfasında.",
            f"🔴 {admin_crit} yetkisiz admin hesabı — Bugün AD'den kaldırılmalı.",
            f"🔴 {patch_crit} cihaz 60+ gün yamasız — Bu hafta zorunlu güncelleme.",
            f"🔴 {eol_count} EoL OS (Win7/2008/8.1) — Upgrade takvimi hazırlanmalı.",
            f"⚡ {n_servers} kritik varlık tespit edildi, Criticality çarpanı uygulandı.",
            f"✅ {n_low} cihaz (%{round(n_low/n_total*100,1) if n_total else 0}) düşük risk kategorisinde.",
        ]
        AI_RISK_TREND   = f"En kritik tehdit bu hafta: Patch uyumsuzluğu ({patch_crit} cihaz) ve yetkisiz admin ({admin_crit} cihaz)."
        AI_SUBJECT_LINE = ""

    # ═══════════════════════════════════════════════════════════
    # 3. GRAFİKLER
    # ═══════════════════════════════════════════════════════════
    log("Grafikler olusturuluyor...")

    def _savefig(fig, name):
        path = os.path.join(TEMP_DIR, name)
        fig.savefig(path, dpi=140, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)
        return path

    def _ax(ax, title, xl="", yl=""):
        ax.set_facecolor("#1C2833")
        ax.set_title(title, color="white", fontsize=11, fontweight="bold", pad=8)
        ax.set_xlabel(xl, color="#AAB7B8", fontsize=9)
        ax.set_ylabel(yl, color="#AAB7B8", fontsize=9)
        ax.tick_params(colors="#AAB7B8", labelsize=8)
        for sp in ax.spines.values():
            sp.set_edgecolor("#2C3E50")

    def pie(vals, labels, colors, title, fname):
        fig, ax = plt.subplots(figsize=(5, 4), facecolor=BG)
        nz = [(v,l,c) for v,l,c in zip(vals,labels,colors) if v > 0]
        if not nz: nz = [(1,"Veri Yok","#555")]
        vs, ls, cs = zip(*nz)
        _, _, auts = ax.pie(vs, labels=None, colors=cs, autopct="%1.1f%%",
                            startangle=140, pctdistance=0.72,
                            wedgeprops=dict(edgecolor="white", linewidth=2.5))
        for at in auts:
            at.set_color("white"); at.set_fontsize(11); at.set_fontweight("bold")
        ax.legend(ls, loc="lower center", ncol=len(ls), framealpha=0,
                  labelcolor="white", fontsize=9, bbox_to_anchor=(0.5, -0.1))
        ax.set_title(title, color="white", fontsize=11, fontweight="bold", pad=12)
        fig.patch.set_facecolor(BG)
        return _savefig(fig, fname)

    def hbar(cats, vals, colors, title, fname):
        h = max(3.2, len(cats) * 0.52)
        fig, ax = plt.subplots(figsize=(6.5, h), facecolor=BG)
        ax.set_facecolor("#1C2833")
        bars = ax.barh(cats, vals, color=colors, edgecolor="none", height=0.56)
        mx = max(vals) if vals else 1
        for bar in bars:
            w = bar.get_width()
            ax.text(w + mx*0.02, bar.get_y()+bar.get_height()/2,
                    f"{int(w)}", va="center", color="white", fontsize=9, fontweight="bold")
        _ax(ax, title, "Cihaz Sayısı")
        ax.set_xlim(0, mx * 1.2)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.grid(axis="x", alpha=0.12, color="white")
        fig.tight_layout()
        return _savefig(fig, fname)

    def hist(series, color, title, xl, fname, bins=14, vlines=None):
        fig, ax = plt.subplots(figsize=(5.2, 3.6), facecolor=BG)
        ax.hist(series.dropna(), bins=bins, color=color, edgecolor=BG, alpha=0.88)
        ax.axvline(series.mean(), color="#F1C40F", lw=2, linestyle="--",
                   label=f"Ort: {series.mean():.1f}")
        if vlines:
            for v, vc, vl in vlines:
                ax.axvline(v, color=vc, lw=1.5, linestyle=":", label=vl)
        ax.legend(framealpha=0, labelcolor="white", fontsize=8)
        _ax(ax, title, xl, "Cihaz Sayısı")
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        fig.tight_layout()
        return _savefig(fig, fname)

    def vbar(cats, vals, colors, title, xl, yl, fname):
        fig, ax = plt.subplots(figsize=(6, 4), facecolor=BG)
        ax.set_facecolor("#1C2833")
        mx = max(vals) if vals else 1
        bars = ax.bar(cats, vals, color=colors, edgecolor="none", width=0.58)
        for bar in bars:
            h2 = bar.get_height()
            ax.text(bar.get_x()+bar.get_width()/2, h2+mx*0.02,
                    f"{int(h2)}", ha="center", color="white", fontsize=9, fontweight="bold")
        _ax(ax, title, xl, yl)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.12, color="white")
        plt.xticks(rotation=25, ha="right")
        fig.tight_layout()
        return _savefig(fig, fname)

    def user_bar(df_u, fname):
        top = df_u.head(12)
        fig, ax = plt.subplots(figsize=(8, 5.5), facecolor=BG)
        ax.set_facecolor("#1C2833")
        cl = ["#C0392B" if v>=50 else "#D35400" if v>=25 else "#1E8449"
              for v in top["Davraniş_Risk_Puani"]]
        bars = ax.barh(top["Kullanıcı"], top["Davraniş_Risk_Puani"],
                       color=cl, edgecolor="none", height=0.6)
        mx = max(top["Davraniş_Risk_Puani"]) if len(top) > 0 else 1
        for bar in bars:
            w = bar.get_width()
            ax.text(w+mx*0.02, bar.get_y()+bar.get_height()/2,
                    f"{w:.0f}", va="center", color="white", fontsize=9, fontweight="bold")
        _ax(ax, "En Riskli Kullanıcılar (Davranış Skoru)", "Risk Puanı")
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.grid(axis="x", alpha=0.12, color="white")
        ax.axvline(50, color="#E74C3C", lw=1.5, linestyle="--", alpha=0.7, label="Yüksek Sınırı (50)")
        ax.axvline(25, color="#F39C12", lw=1.5, linestyle=":", alpha=0.7, label="Orta Sınırı (25)")
        ax.legend(framealpha=0, labelcolor="white", fontsize=8)
        fig.tight_layout()
        return _savefig(fig, fname)

    def scatter(df_s, fname):
        fig, ax = plt.subplots(figsize=(6.5, 4.5), facecolor=BG)
        ax.set_facecolor("#1C2833")
        cm = df_s["Seviye"].map({"YUKSEK":"#E74C3C","ORTA":"#F39C12","DUSUK":"#27AE60"}).fillna("#888")
        ax.scatter(df_s["Yamasız Gün"], df_s["Final_Risk_Skoru"],
                   c=cm, alpha=0.8, s=45, edgecolors="none")
        ax.axhline(50, color="#E74C3C", lw=1.5, linestyle="--", alpha=0.7, label="Yüksek Sınırı (50)")
        ax.axvline(60, color="#F39C12", lw=1.5, linestyle="--", alpha=0.7, label="Kritik Patch Sınırı (60g)")
        ax.legend(framealpha=0, labelcolor="white", fontsize=9)
        _ax(ax, "Final Risk Skoru vs Yamasız Gün (Criticality Uygulandı)", "Yamasız Gün", "Final Risk Skoru (0-100)")
        ax.annotate("← Güvenli Bölge", xy=(10, 10), color="#27AE60", fontsize=8)
        ax.annotate("Tehlikeli Bölge →", xy=(80, 70), color="#E74C3C", fontsize=8)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        fig.tight_layout()
        return _savefig(fig, fname)

    def sev_bar(fname):
        """Seviye bazlı ortalama metrikleri karşılaştıran grouped bar"""
        sevs = ["DUSUK","ORTA","YUKSEK"]
        lbls = ["Düşük Risk","Orta Risk","Yüksek Risk"]
        cols_m = ["#27AE60","#F39C12","#E74C3C"]
        metrics = ["Final_Risk_Skoru","Yamasız Gün","Offline Gün"]
        metric_lbls = ["Ort. Final Risk Skoru","Ort. Yamasız Gün","Ort. Offline Gün"]

        x = np.arange(len(metrics))
        width = 0.25
        fig, ax = plt.subplots(figsize=(7, 4.5), facecolor=BG)
        ax.set_facecolor("#1C2833")
        for i, (sev, lbl, col) in enumerate(zip(sevs, lbls, cols_m)):
            sub = df[df["Seviye"] == sev]
            vals_s = [sub[m].mean() if len(sub) > 0 else 0 for m in metrics]
            bars2 = ax.bar(x + i*width, vals_s, width, label=lbl, color=col, alpha=0.85, edgecolor="none")
            for bar in bars2:
                h2 = bar.get_height()
                ax.text(bar.get_x()+bar.get_width()/2, h2+1,
                        f"{h2:.0f}", ha="center", color="white", fontsize=8)
        ax.set_xticks(x + width)
        ax.set_xticklabels(metric_lbls, color="#AAB7B8")
        ax.legend(framealpha=0, labelcolor="white", fontsize=9)
        _ax(ax, "Risk Seviyelerine Göre Ortalama Metrikler", "", "Ortalama Değer")
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.grid(axis="y", alpha=0.12, color="white")
        fig.tight_layout()
        return _savefig(fig, fname)

    def criticality_chart(fname):
        """Cihaz tipi dağılımı + ortalama final skor"""
        ct = df.groupby("Cihaz_Tipi").agg(
            Sayi=("Final_Risk_Skoru","count"),
            Ort_Skor=("Final_Risk_Skoru","mean"),
        ).sort_values("Ort_Skor", ascending=False)

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4.2), facecolor=BG)
        colors = ["#E74C3C","#C0392B","#D35400","#E67E22","#2980B9","#3498DB","#27AE60"]

        # Sol: Cihaz sayısı
        ax1.barh(ct.index, ct["Sayi"],
                 color=colors[:len(ct)], edgecolor="none", height=0.55)
        ax1.set_facecolor("#1C2833")
        mx = ct["Sayi"].max() if len(ct) > 0 else 1
        for i, (_, row) in enumerate(ct.iterrows()):
            ax1.text(row["Sayi"]+mx*0.02, i, f"{int(row['Sayi'])}", va="center",
                     color="white", fontsize=9, fontweight="bold")
        _ax(ax1, "Cihaz Tipi Dağılımı", "Cihaz Sayısı")
        ax1.spines["top"].set_visible(False); ax1.spines["right"].set_visible(False)
        ax1.grid(axis="x", alpha=0.12, color="white")

        # Sağ: Ortalama final skor (sunucular çok daha yüksek çıkacak)
        bar_cols = ["#E74C3C" if s>=50 else "#F39C12" if s>=25 else "#27AE60"
                    for s in ct["Ort_Skor"]]
        ax2.barh(ct.index, ct["Ort_Skor"],
                 color=bar_cols, edgecolor="none", height=0.55)
        ax2.set_facecolor("#1C2833")
        mx2 = ct["Ort_Skor"].max() if len(ct) > 0 else 1
        for i, (_, row) in enumerate(ct.iterrows()):
            ax2.text(row["Ort_Skor"]+mx2*0.02, i, f"{row['Ort_Skor']:.1f}", va="center",
                     color="white", fontsize=9, fontweight="bold")
        ax2.axvline(50, color="#E74C3C", lw=1.5, linestyle="--", alpha=0.7, label="Yüksek Sınırı (50)")
        ax2.legend(framealpha=0, labelcolor="white", fontsize=8)
        _ax(ax2, "Cihaz Tipine Göre Ort. Final Risk Skoru", "Ort. Final Risk Skoru")
        ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)
        ax2.grid(axis="x", alpha=0.12, color="white")

        fig.suptitle("Asset Criticality Analizi  |  Sunucular öncelikli izlenmeli",
                     color="white", fontsize=11, fontweight="bold")
        fig.tight_layout(); fig.patch.set_facecolor(BG)
        return _savefig(fig, fname)

    def dashboard_panel(fname):
        fig, axes = plt.subplots(1, 3, figsize=(13, 4.2), facecolor=BG)
        fig.suptitle(f"IT Siber Güvenlik Yönetici Özeti  |  {TARIH_H}",
                     color="white", fontsize=12, fontweight="bold")
        pv=[n_high,n_medium,n_low]; pc=["#E74C3C","#F39C12","#27AE60"]
        pl=[f"Yüksek\n{n_high}",f"Orta\n{n_medium}",f"Düşük\n{n_low}"]
        nz=[(pv[i],pc[i],pl[i]) for i in range(3) if pv[i]>0]
        if nz:
            vs,cs,ls=zip(*nz)
            _,_,auts=axes[0].pie(vs,colors=cs,autopct="%1.0f%%",startangle=90,
                                  wedgeprops=dict(edgecolor="white",lw=2.5))
            for at in auts: at.set_color("white"); at.set_fontsize(11); at.set_fontweight("bold")
            axes[0].legend(ls,loc="lower center",ncol=3,framealpha=0,
                           labelcolor="white",fontsize=8,bbox_to_anchor=(0.5,-0.2))
        axes[0].set_title("Risk Dağılımı",color="white",fontsize=10,fontweight="bold")
        axes[0].set_facecolor("#1C2833")

        td=threat_df.head(7)
        bc=["#E74C3C","#C0392B","#D35400","#E67E22","#8E44AD","#2980B9","#16A085"][:len(td)]
        axes[1].barh(td["Tehdit"],td["Sayi"],color=bc,edgecolor="none")
        axes[1].set_facecolor("#1C2833")
        axes[1].set_title("Tehdit Türleri (Cihaz Sayısı)",color="white",fontsize=10,fontweight="bold")
        axes[1].tick_params(colors="#AAB7B8",labelsize=7)
        axes[1].spines["top"].set_visible(False); axes[1].spines["right"].set_visible(False)
        axes[1].grid(axis="x",alpha=0.12,color="white")
        for sp in ["left","bottom"]: axes[1].spines[sp].set_edgecolor("#2C3E50")

        axes[2].hist(df["Risk Skoru"].dropna(),bins=14,color="#3498DB",edgecolor=BG,alpha=0.88)
        axes[2].axvline(50,color="#E74C3C",lw=2,linestyle="--",label="Yüksek (50)")
        axes[2].axvline(25,color="#F39C12",lw=1.5,linestyle=":",label="Orta (25)")
        axes[2].legend(framealpha=0,labelcolor="white",fontsize=7)
        axes[2].set_facecolor("#1C2833")
        axes[2].set_title("Risk Skoru Dağılımı",color="white",fontsize=10,fontweight="bold")
        axes[2].tick_params(colors="#AAB7B8",labelsize=8)
        axes[2].spines["top"].set_visible(False); axes[2].spines["right"].set_visible(False)
        axes[2].grid(axis="y",alpha=0.12,color="white")
        for sp in ["left","bottom"]: axes[2].spines[sp].set_edgecolor("#2C3E50")
        for ax in axes: ax.set_facecolor("#1C2833")
        fig.tight_layout(); fig.patch.set_facecolor(BG)
        return _savefig(fig, fname)

    # Üret
    pie_path        = pie([n_high,n_medium,n_low],
                          [f"Yüksek Risk ({n_high})",f"Orta Risk ({n_medium})",f"Düşük Risk ({n_low})"],
                          ["#E74C3C","#F39C12","#27AE60"],"Risk Seviyesi Dağılımı","pie_risk.png")
    pie_mail_path   = pie([n_high,n_medium,n_low],
                          [f"Yüksek ({n_high})",f"Orta ({n_medium})",f"Düşük ({n_low})"],
                          ["#E74C3C","#F39C12","#27AE60"],"Risk Dağılımı","pie_mail.png")
    hist_score_path = hist(df["Final_Risk_Skoru"],"#E74C3C","Risk Skoru Dağılımı (Criticality Uygulandı)",
                           "Final Risk Skoru (0=Güvenli, 100=Kritik)","hist_score.png",
                           vlines=[(50,"#E74C3C","Yüksek Sınırı (50)"),(25,"#F39C12","Orta Sınırı (25)")])
    hist_off_path   = hist(df["Offline Gün"],"#F39C12","Offline Süre Dağılımı",
                           "Offline Gün (0=Bugün Görüldü)","hist_offline.png",
                           vlines=[(60,"#E74C3C","Kritik Sınırı (60 gün)")])
    hist_patch_path = hist(df["Yamasız Gün"],"#3498DB","Yamasız Süre Dağılımı",
                           "Yamasız Gün (0=Bugün Yamalı)","hist_patch.png",
                           vlines=[(60,"#E74C3C","Kritik Sınırı (60 gün)")])
    hist_disk_path  = hist(df["% Boş"],"#27AE60","Boş Disk Alanı Dağılımı",
                           "Boş Disk % (100=Tamamen Boş)","hist_disk.png",
                           vlines=[(10,"#E74C3C","Kritik (<10% Boş)")])
    tc = ["#1ABC9C","#16A085","#2980B9","#8E44AD","#D35400","#E67E22","#C0392B","#E74C3C","#922B21","#7B241C"]
    bar_threat_path = hbar(threat_df["Tehdit"].tolist(), threat_df["Sayi"].tolist(),
                           tc[:len(threat_df)], "Tehdit Türüne Göre Etkilenen Cihaz Sayısı","bar_threat.png")
    os_counts       = df["Sistem"].value_counts().head(7)
    bar_os_path     = vbar(os_counts.index.tolist(), os_counts.values.tolist(),
                           ["#3498DB","#E74C3C","#F39C12","#27AE60","#8E44AD","#E67E22","#1ABC9C"][:len(os_counts)],
                           "İşletim Sistemi Dağılımı","İşletim Sistemi","Cihaz Sayısı","bar_os.png")
    bar_user_path   = user_bar(user_risk, "bar_users.png")
    scatter_path    = scatter(df, "scatter.png")
    sev_bar_path    = sev_bar("sev_bar.png")
    crit_chart_path = criticality_chart("crit_chart.png")
    dash_path       = dashboard_panel("dashboard.png")

    log("Tüm grafikler hazır.")

    # ═══════════════════════════════════════════════════════════
    # 4. OPENPYXL YARDIMCILARI
    # ═══════════════════════════════════════════════════════════
    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, name="Calibri"):
        return Font(bold=bold, color=color, size=size, name=name)
    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _border(color="CCCCCC"):
        s = Side(border_style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)
    def _title(ws, rng, text, bg=CD, fg=CW, sz=13):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value=text; c.fill=_fill(bg); c.font=_font(True,fg,sz); c.alignment=_align()
    def _img(ws, path, anchor, w=385, h=268):
        if not os.path.exists(path): return
        img = XLImage(path); img.width=w; img.height=h
        ws.add_image(img, anchor)
    def _kpi(ws, row, items):
        # items: [(col, label, value, bg), ...]
        for col, lbl, val, bg in items:
            ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
            ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
            lc = ws.cell(row=row,   column=col, value=lbl)
            vc = ws.cell(row=row+1, column=col, value=val)
            lc.fill=_fill(bg); lc.font=_font(True,CW,9);  lc.alignment=_align()
            vc.fill=_fill(bg); vc.font=_font(True,CW,17); vc.alignment=_align()
        ws.row_dimensions[row].height=22; ws.row_dimensions[row+1].height=40

    def _note(ws, rng, txt, bg, fg):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value=txt; c.fill=_fill(bg); c.font=_font(False,fg,9)
        c.alignment=_align(h="left", wrap=True)

    def _sev(sev):
        if sev=="YUKSEK": return "FFD7D7","C0392B","YÜKSEK"
        if sev=="ORTA":   return "FFF3CD","D35400","ORTA"
        return "D4EDDA","1E8449","DÜŞÜK"

    LANSWEEPER_BASE = "http://LANSWEEPER_HOST:81"   # Lansweeper adresi

    def _lsw_url(asset_id=None, asset_name=None) -> str:
        """Sadece AssetID ile direkt link üret. İsim tabanlı link üretme."""
        if asset_id is not None:
            try:
                aid_str = str(asset_id).strip()
                if aid_str in ("", "nan", "None", "NaN", "0", "0.0"):
                    return ""
                aid = int(float(aid_str))
                if aid > 0:
                    return f"{LANSWEEPER_BASE}/asset.aspx?AssetID={aid}"
            except Exception:
                pass
        return ""

    def _table(ws, start_row, start_col, tdf, hdr_map, hdr_bg,
               sev_col=None, cond_cols=None, row_ht=None):
        cols = list(tdf.columns)
        asset_id_col   = "AssetID"   if "AssetID"   in tdf.columns else None
        asset_name_col = "AssetName" if "AssetName" in tdf.columns else None

        for ci, col in enumerate(cols, start_col):
            c = ws.cell(row=start_row, column=ci, value=hdr_map.get(col, col))
            c.fill=_fill(hdr_bg); c.font=_font(True,CW,9)
            c.alignment=_align(wrap=True); c.border=_border()
        ws.row_dimensions[start_row].height = 36
        for ri, (_, rd) in enumerate(tdf.iterrows(), start_row+1):
            sev = str(rd.get(sev_col,"")) if sev_col else "DUSUK"
            rbg, sbg, slbl = _sev(sev)
            ht = row_ht if row_ht else 18
            ws.row_dimensions[ri].height = ht
            # Lansweeper URL — SADECE AssetID ile, isim araması yapma
            lsw_url = _lsw_url(asset_id=rd.get(asset_id_col) if asset_id_col else None)

            for ci, col in enumerate(cols, start_col):
                val = rd.get(col)
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill=_fill(rbg); c.font=_font(False,"000000",9)
                c.alignment=_align(h="center", wrap=True); c.border=_border()
                # Hem AssetID hem AssetName sütununa link ekle
                if col in (asset_id_col, asset_name_col) and lsw_url:
                    c.hyperlink = lsw_url
                    c.font = _font(False,"0563C1",9)
                if col == sev_col:
                    c.value=slbl; c.fill=_fill(sbg); c.font=_font(True,CW,9)
                if cond_cols and col in cond_cols and val is not None:
                    try:
                        fv = float(val)
                        for lo, hi, cbg, cfg in cond_cols[col]:
                            if lo <= fv < hi:
                                c.fill=_fill(cbg); c.font=_font(True,cfg,9); break
                    except: pass

    # ═══════════════════════════════════════════════════════════
    # 5. EXCEL – 8 SHEET
    # ═══════════════════════════════════════════════════════════
    log("Excel raporu oluşturuluyor (11 sheet)...")
    wb = Workbook(); wb.remove(wb.active)

    # ── SHEET 1 – Genel Özet ──────────────────────────────────
    ws1 = wb.create_sheet("01 Genel Ozet")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "B17"
    _title(ws1,"B2:P3",
           f"IT SİBER GÜVENLİK HAFTALIK RAPORU  |  {TARIH_H}  |  Toplam {n_total} Cihaz",
           "12181F",CW,14)
    ws1.row_dimensions[2].height=30; ws1.row_dimensions[3].height=30

    _kpi(ws1,5,[
        (2,"Toplam Cihaz",n_total,"2C3E50"),
        (4,"🔴 Yüksek Risk",n_high,"C0392B"),
        (6,"🟠 Orta Risk",n_medium,"D35400"),
        (8,"🟢 Düşük Risk",n_low,"1E8449"),
        (10,"⚠ Patch >60g",patch_crit,"6C3483"),
        (12,"📴 Offline >60g",off_crit,"117A65"),
        (14,"💀 EoL OS",eol_count,"784212"),
        (16,"🛡 CVE Kritik",cve_kritik,"7B241C"),
    ])

    _img(ws1, pie_path, "B8", 350, 268)

    ws1.merge_cells("H8:P8")
    c=ws1["H8"]; c.value="ÖZET & ACİL UYARILAR"
    c.fill=_fill("12181F"); c.font=_font(True,CW,12); c.alignment=_align()

    box_items = [
        (9,"H","P","YÜKSEK RİSKLİ CİHAZLAR → Bu hafta acil inceleme gerektirir! BT ekibi önceliklendirmeli.","FFCCCC","7B241C"),
        (10,"H","P",f"{admin_crit} cihazda yetkisiz admin → AD'den derhal kaldırılmalı, log kaydı alınmalı!","FFEECC","7D3C00"),
        (11,"H","P",f"{patch_crit} cihaz 60+ gün yamasız → Fidye yazılımı saldırısına açık kapı! Acil yamala.","FFEECC","7D3C00"),
        (12,"H","P",f"{eol_count} cihaz desteksiz OS kullanıyor → Win 7/2008/8.1'e artık güvenlik yamasi çıkmıyor!","FFCCCC","7B241C"),
        (13,"H","P",f"{off_crit} cihaz 60+ gün offline → Zombi cihaz. Kimin kullandığı bilinmiyor!","D5F5E3","1E8449"),
        (14,"H","P",f"Ort. risk skoru: {avg_score}/100 | Yetkisiz admin: {admin_crit} | Shadow IT: {shadow_crit} cihaz","D6EAF8","1A5276"),
    ]
    for rw,sc,ec,txt,bg,fg in box_items:
        ws1.merge_cells(f"{sc}{rw}:{ec}{rw}")
        cell=ws1[f"{sc}{rw}"]; cell.value=txt; cell.fill=_fill(bg)
        cell.font=_font(False,fg,10); cell.alignment=_align(h="left",wrap=True)
        ws1.row_dimensions[rw].height=22

    # Tam envanter tablosu
    inv_cols=["AssetID","AssetName","Kullanıcı","Durum","IPAddress","Sistem","Cihaz_Tipi",
              "% Boş","Offline Gün","Yamasız Gün","Risk Skoru","Final_Risk_Skoru","Risk Analizi","Seviye"]
    inv_df=df[[c for c in inv_cols if c in df.columns]].copy()
    inv_df=inv_df.sort_values("Final_Risk_Skoru",ascending=False)
    inv_hdr={
        "AssetID":"Cihaz ID","AssetName":"Cihaz Adı",
        "Kullanıcı":"Kullanıcı","Durum":"Durum","IPAddress":"IP Adresi",
        "Sistem":"İşletim Sistemi",
        "Cihaz_Tipi":"Cihaz Tipi\n(Criticality)",
        "% Boş":"Boş Disk (%)",
        "Offline Gün":"Offline (Gün)","Yamasız Gün":"Yamasız (Gün)",
        "Risk Skoru":"Lansweeper\nRisk Skoru",
        "Final_Risk_Skoru":"Final Risk Skoru\n(Criticality Uygulandı ⭐)",
        "Risk Analizi":"Risk Analizi","Seviye":"Risk Seviyesi",
    }
    T1=17
    ws1.cell(row=T1-1,column=2,
             value=f"TAM CİHAZ ENVANTERİ – Final Riske Göre Sıralı ({n_total} Cihaz)  ⭐=Criticality çarpanı uygulandı").font=_font(True,CD,11)
    _table(ws1,T1,2,inv_df,inv_hdr,CD,"Seviye",
           cond_cols={
               "Final_Risk_Skoru":[(75,101,"FFCCCC","7B241C"),(50,75,"FFD7D7","990000"),(25,50,"FFF3CD","7D3C00")],
               "Risk Skoru":[(75,101,"FFCCCC","7B241C"),(50,75,"FFD7D7","990000"),(25,50,"FFF3CD","7D3C00")],
               "% Boş":[(-9999,10,"FFCCCC","990000")],
               "Offline Gün":[(60,9999,"FFCCCC","990000")],
               "Yamasız Gün":[(60,9999,"FFF3CD","7D3C00")],
           })
    for ci,w in enumerate([4,10,13,14,10,14,16,22,10,16,16,18,45,13],1):
        ws1.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 1 OK")

    # ── SHEET 2 – Teknik Metrikler ────────────────────────────
    ws2 = wb.create_sheet("02 Teknik Metrikler")
    ws2.sheet_view.showGridLines = False
    _title(ws2,"B2:M3","TEKNİK RİSK METRİKLERİ  |  Disk · Offline · Patch · Risk Skoru","1A5276",CW,13)

    _kpi(ws2,5,[
        (2,"Ort. Risk Skoru",f"{avg_score}","1A5276"),
        (4,"Maks. Risk Skoru",f"{max_score}","922B21"),
        (6,"Ort. Yamasız Gün",f"{df['Yamasız Gün'].mean():.1f}","784212"),
        (8,"Ort. Offline Gün",f"{df['Offline Gün'].mean():.1f}","145A32"),
        (10,"Kritik Disk (<10%)",disk_crit,"6C3483"),
        (12,"Patch Kritik (>60g)",patch_crit,"C0392B"),
    ])

    # 4 histogram - 2x2 grid - aralarında boşluk bırak
    _img(ws2, hist_score_path, "B8",  382, 268)
    _img(ws2, hist_off_path,   "H8",  382, 268)
    # Her bir bloğun altına yorum satırı (satır 22-23)
    _note(ws2,"B22:G23","→ 50 üzeri skor: Acil müdahale! Bu cihazlar fidye yazılımı ve APT saldırılarına karşı savunmasız.","FFCCCC","990000")
    _note(ws2,"H22:M23","→ 60+ gün offline: Zombi cihaz riski. Bu cihazlar ağda kim tarafından kullanılıyor bilinmiyor!","FFF3CD","7D3C00")
    # 2. satır grafikler (satır 25'ten başla)
    _img(ws2, hist_patch_path, "B25", 382, 268)
    _img(ws2, hist_disk_path,  "H25", 382, 268)
    _note(ws2,"B39:G40","→ 60+ gün yamasız: Microsoft'un yayımladığı kritik CVE yamalarından yoksun cihazlar. Hemen patch!","FFF3CD","7D3C00")
    _note(ws2,"H39:M40","→ Boş disk < %10: Sistem yavaşlaması, log kaybı ve servis çökmesi riski. Temizleme yapılmalı.","FFD7D7","990000")
    # Scatter
    _img(ws2, scatter_path,    "B42", 510, 320)
    _note(ws2,"B57:M58","OKUMA KILAVUZU: Sağ-üst köşedeki noktalar en tehlikeli cihazlardır (hem uzun süredir yamasız hem yüksek riskli). Bu noktaları öncelikle ele alın!","FFCCCC","990000")
    # Seviye grouped bar
    _img(ws2, sev_bar_path,    "B60", 480, 300)
    _note(ws2,"B76:M77","Bu grafik, yüksek riskli grupta yamasız gün ve offline gün ortalamalarının neden çok daha yüksek olduğunu göstermektedir.","D6EAF8","1A5276")

    tech_df=df[["AssetName","Kullanıcı","% Boş","Offline Gün","Yamasız Gün","Risk Skoru","Seviye"]].copy()
    T2=80
    ws2.cell(row=T2-1,column=2,value="CİHAZ TEKNİK METRİK TABLOSU (Tüm Cihazlar)").font=_font(True,CD,11)
    _table(ws2,T2,2,tech_df.sort_values("Risk Skoru",ascending=False),{
        "AssetName":"Cihaz Adı",
        "Kullanıcı":"Kullanıcı",
        "% Boş":"Boş Disk (%)\n(Kritik: <10 = Disk Dolu!)",
        "Offline Gün":"Offline Gün\n(Kritik: >60 = Zombi Cihaz!)",
        "Yamasız Gün":"Yamasız Gün\n(Kritik: >60 = CVE Açığı!)",
        "Risk Skoru":"Risk Skoru\n(0=Güvenli | 50=Yüksek | 100=Kritik)",
        "Seviye":"Risk Seviyesi",
    },"1A5276","Seviye",
    cond_cols={
        "% Boş":[(-9999,10,"FFCCCC","990000")],
        "Offline Gün":[(60,9999,"FFCCCC","990000")],
        "Yamasız Gün":[(60,9999,"FFF3CD","7D3C00")],
        "Risk Skoru":[(75,101,"FFCCCC","7B241C"),(50,75,"FFD7D7","990000"),(25,50,"FFF3CD","7D3C00")],
    })
    for ci,w in enumerate([4,16,15,16,18,18,22,14],1): ws2.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 2 OK")

    # ── SHEET 3 – Güvenlik Bulguları ─────────────────────────
    ws3 = wb.create_sheet("03 Guvenlik Bulgulari")
    ws3.sheet_view.showGridLines = False
    _title(ws3,"B2:L3","GÜVENLİK BULGULARI  |  Admin · Paylaşım · Shadow IT · EoL OS","6C3483",CW,13)
    _kpi(ws3,5,[
        (2,"Yetkisiz Admin",admin_crit,"922B21"),
        (4,"Riskli Klasör",share_crit,"784212"),
        (6,"Shadow IT",shadow_crit,"1A5276"),
        (8,"EoL OS",eol_count,"145A32"),
    ])
    _img(ws3, bar_threat_path, "B8",  515, 325)
    _img(ws3, bar_os_path,     "J8",  440, 325)

    acil=[
        ("B24","L24","YÜKSEK ÖNCELİKLİ GÜVENLİK EYLEMLERİ – Bu Hafta Tamamlanmalı!","922B21",CW,True,22),
        ("B25","L25",f"1. Yetkisiz {admin_crit} admin hesabı → Bu GÜN AD'den kaldırılmalı, log alınmalı, yöneticiye bildirilmeli!","FFD7D7","990000",False,18),
        ("B26","L26",f"2. {share_crit} riskli klasör paylaşımı → Herkese açık (Everyone/Domain Users) erişim derhal kapatılmalı.","FFF3CD","7D3C00",False,18),
        ("B27","L27",f"3. {shadow_crit} cihazda TeamViewer/AnyDesk/Torrent/VPN → GPO script ile uzaktan kaldırılmalı.","FFD7D7","990000",False,18),
        ("B28","L28",f"4. {eol_count} cihaz desteksiz OS (Win7/2008/8.1) → Artık güvenlik yaması çıkmıyor! Upgrade takvimi hazırlanmalı.","FFD7D7","990000",False,18),
        ("B29","L29","5. SEP/Antivirüs eksik cihazlara → SCCM/GPO ile deployment yapılmalı, tamamlanma raporu alınmalı.","D5F5E3","1E8449",False,18),
    ]
    for s,e,txt,bg,fg,bold,ht in acil:
        ws3.merge_cells(f"{s}:{e}")
        c=ws3[s]; c.value=txt; c.fill=_fill(bg); c.font=_font(bold,fg,10)
        c.alignment=_align(h="left",wrap=True); ws3.row_dimensions[int(s[1:])].height=ht

    sec_cols=["AssetName","Kullanıcı","Kural Dışı Adminler (İsim ve Ünvan)",
              "Riskli Paylaşılan Klasörler","Tespit Edilen Şüpheli Yazılımlar","Seviye"]
    sec_df=df[[c for c in sec_cols if c in df.columns]].copy()
    has_find=(
        sec_df.get("Kural Dışı Adminler (İsim ve Ünvan)",pd.Series()).notna() |
        sec_df.get("Riskli Paylaşılan Klasörler",pd.Series()).notna() |
        sec_df.get("Tespit Edilen Şüpheli Yazılımlar",pd.Series()).notna()
    )
    sec_filt=sec_df[has_find].reset_index(drop=True)
    T3=32
    ws3.cell(row=T3-1,column=2,
             value=f"GÜVENLİK BULGULARI – YALNIZCA BULGULU CİHAZLAR ({len(sec_filt)} adet)").font=_font(True,CD,11)
    _table(ws3,T3,2,sec_filt,{
        "AssetName":"Cihaz Adı",
        "Kullanıcı":"Kullanıcı",
        "Kural Dışı Adminler (İsim ve Ünvan)":"Kural Dışı Admin\n(İsim & Unvan | Açıklama: IT onayı olmayan admin yetkisi)",
        "Riskli Paylaşılan Klasörler":"Riskli Klasörler\n(Herkese Açık Yazma/Tam Yetki = Veri Sızıntısı Riski!)",
        "Tespit Edilen Şüpheli Yazılımlar":"Yasak/Şüpheli Yazılımlar\n(Uzaktan Erişim, P2P, Hacking Araçları)",
        "Seviye":"Risk Seviyesi",
    },"6C3483","Seviye")
    for ci,w in enumerate([4,16,15,28,25,28,14],1): ws3.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 3 OK")

    # ── SHEET 4 – Kullanıcı Risk (DETAYLI) ───────────────────
    ws4 = wb.create_sheet("04 Kullanici Risk")
    ws4.sheet_view.showGridLines = False
    _title(ws4,"B2:P3",
           "KULLANICI DAVRANIŞ RİSK ANALİZİ  |  Kim, Kaç Cihazda, Ne Kadar Tehdit Oluşturuyor?",
           "117A65",CW,13)

    _kpi(ws4,5,[
        (2,"Toplam Kullanıcı",int(user_risk["Kullanıcı"].nunique() if "Kullanıcı" in user_risk.columns else len(user_risk)),"117A65"),
        (4,"Yüksek Riskli Kull.",int((user_risk["Davraniş_Risk_Puani"]>=50).sum()),"C0392B"),
        (6,"Orta Riskli Kull.",int(((user_risk["Davraniş_Risk_Puani"]>=25)&(user_risk["Davraniş_Risk_Puani"]<50)).sum()),"D35400"),
        (8,"En Yüksek Puan",f"{user_risk['Davraniş_Risk_Puani'].max():.0f}","922B21"),
    ])

    _img(ws4, bar_user_path, "B8", 575, 395)

    # Formül açıklaması - daha geniş
    ws4.merge_cells("J8:P8")
    ws4["J8"].value="PUAN HESAPLAMA FORMÜLÜ – NASIL ÇALIŞIR?"
    ws4["J8"].fill=_fill("117A65"); ws4["J8"].font=_font(True,CW,11); ws4["J8"].alignment=_align()

    formula=[
        ("J9","P9","DAVRANIJ RİSK PUANI = (aşağıdaki tüm puanların toplamı)","D6EAF8","1A5276",True),
        ("J10","P10","+ Ortalama Cihaz Risk Skoru (0-100)\n  → Kullanıcıya ait tüm cihazların ortalama tehlike düzeyi","EAF7FB","1A5276",False),
        ("J11","P11","+ Yetkisiz Admin Sayısı × 10 puan\n  → Her bir kural dışı admin hesabı 10 puan ekler","FFF3CD","7D3C00",False),
        ("J12","P12","+ 20 puan (eğer kullanıcıya ait herhangi bir cihaz 60+ gün yamasızsa)\n  → Tek yamasız cihaz bile tüm kullanıcıyı riskli yapar","FFF3CD","7D3C00",False),
        ("J13","P13","+ Disk Hata Sayısı × 5 puan\n  → Kritik disk doluluk/arızası olan her cihaz 5 puan ekler","FFEECC","784212",False),
        ("J14","P14","+ WU Servisi Kapalı Cihaz × 5 puan\n  → Windows Update servisi kapalı her cihaz 5 puan ekler","FFEECC","784212",False),
        ("J15","P15","━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━","EAFAF1","117A65",False),
        ("J16","P16","TOPLAM 50+ PUAN  →  YÜKSEK RİSKLİ KULLANICI\n  ⚠ Derhal incelenmeli! Bu kullanıcının tüm cihazları gözden geçirilmeli.","FFD7D7","990000",True),
        ("J17","P17","TOPLAM 25-49 PUAN  →  ORTA RİSKLİ KULLANICI\n  → Bu hafta takip edilmeli, eksiklikler tamamlanmalı.","FFF3CD","7D3C00",True),
        ("J18","P18","TOPLAM 0-24 PUAN  →  DÜŞÜK RİSKLİ KULLANICI\n  → Rutin izlem yeterli.","D4EDDA","1E8449",True),
    ]
    for s,e,txt,bg,fg,bold in formula:
        ws4.merge_cells(f"{s}:{e}")
        c=ws4[s]; c.value=txt; c.fill=_fill(bg)
        c.font=_font(bold,fg,9); c.alignment=_align(h="left",wrap=True); c.border=_border("AAAAAA")
        rn=int(s[1:])
        ws4.row_dimensions[rn].height=40 if "\n" in txt else 22

    # Kullanıcı detay tablosu
    user_tbl=user_risk.copy()
    # Sütun adlarını detaylı yap
    user_tbl = user_tbl.rename(columns={
        "Kullanıcı":           "Kullanıcı\n(Active Directory Hesabı)",
        "Cihaz_Sayisi":        "Toplam Cihaz\n(Sorumlu Olduğu Cihaz Sayısı)",
        "Ort_Risk_Skoru":      "Ort. Risk Skoru\n(Cihazlarının Ortalama Tehlike Puanı 0-100)",
        "En_Yuksek_Skor":      "En Yüksek Skor\n(En Tehlikeli Cihazının Puanı)",
        "Yuksek_Riskli":       "Yüksek Riskli\nCihaz Sayısı (Skor ≥50)",
        "Max_Yamasiz_Gun":     "En Uzun Yamasız\nGün (En Kötü Cihaz)",
        "Toplam_Admin":        "Yetkisiz Admin\nSayısı (Kural Dışı Yetki)",
        "Disk_Hata_Sayi":      "Disk Arızası\n(Kritik Doluluk Cihaz Sayısı)",
        "WU_Kapali_Sayi":      "WU Kapalı\n(Win Update Servisi Durmuş Cihaz)",
        "Davraniş_Risk_Puani": "DAVRANIŞ RİSK PUANI\n(Toplam = Bu Kullanıcının Risk Skoru)",
    })
    T4=22
    ws4.cell(row=T4-1,column=2,
             value="KULLANICI DAVRANIŞ RİSK SKORU TABLOSU (İlk 15 – En Riskli Kullanıcılar)").font=_font(True,CD,11)
    cols4=list(user_tbl.columns)
    for ci,col in enumerate(cols4,2):
        c=ws4.cell(row=T4,column=ci,value=col)
        c.fill=_fill("117A65"); c.font=_font(True,CW,9)
        c.alignment=_align(wrap=True); c.border=_border()
    ws4.row_dimensions[T4].height=50
    for ri,(_,rd) in enumerate(user_tbl.iterrows(),T4+1):
        puan=rd.iloc[-1]
        rbg="FFD7D7" if puan>=50 else ("FFF3CD" if puan>=25 else "D4EDDA")
        ws4.row_dimensions[ri].height=22
        for ci,val in enumerate(rd,2):
            c=ws4.cell(row=ri,column=ci,value=round(val,1) if isinstance(val,float) else val)
            c.fill=_fill(rbg); c.font=_font(False,"000000",9)
            c.alignment=_align(h="center"); c.border=_border()
        # Son sütun (puan) vurgula
        pc=ws4.cell(row=ri,column=2+len(cols4)-1)
        if puan>=50:   pc.fill=_fill("C0392B"); pc.font=_font(True,CW,11)
        elif puan>=25: pc.fill=_fill("D35400"); pc.font=_font(True,CW,11)
        else:          pc.fill=_fill("1E8449"); pc.font=_font(True,CW,11)
    for ci,w in enumerate([4,24,18,20,18,14,18,14,14,14,22],1):
        ws4.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 4 OK")

    # ── SHEET 5 – OS & EoL ───────────────────────────────────
    ws5=wb.create_sheet("05 OS ve EoL Analizi")
    ws5.sheet_view.showGridLines=False
    _title(ws5,"B2:M3","İŞLETİM SİSTEMİ & EOL RİSK ANALİZİ  |  Artık Güvenlik Yaması Çıkmayan Cihazlar","784212",CW,13)
    eol_df=df[df["Sistem"].str.contains("Win 7|2008|8.1|XP|2012",na=False)][
        ["AssetName","Kullanıcı","Sistem","Offline Gün","Risk Skoru","Seviye"]].copy()
    _kpi(ws5,5,[
        (2,"Toplam EoL Cihaz",eol_count,"784212"),
        (4,"Win 7 (2020 EOL)",int(df["Sistem"].str.contains("Win 7",na=False).sum()),"922B21"),
        (6,"Server 2008 (2020 EOL)",int(df["Sistem"].str.contains("2008",na=False).sum()),"C0392B"),
        (8,"Win 8.1 (2023 EOL)",int(df["Sistem"].str.contains("8.1",na=False).sum()),"D35400"),
    ])
    _img(ws5,bar_os_path,"B8",515,320)
    _img(ws5,scatter_path,"J8",448,320)
    ws5.merge_cells("B22:L22")
    c=ws5["B22"]; c.value="⛔ UYARI: Bu cihazlara artık Microsoft tarafından GÜVENLİK YAMASI yayımlanmamaktadır!"
    c.fill=_fill("922B21"); c.font=_font(True,CW,12); c.alignment=_align()
    ws5.merge_cells("B23:L25")
    c=ws5["B23"]
    c.value=("Windows 7 destek sonu: 14 Ocak 2020  |  Windows Server 2008 R2 destek sonu: 14 Ocak 2020\n"
             "Windows 8.1 destek sonu: 10 Ocak 2023  |  Windows Server 2012 destek sonu: 10 Ekim 2023\n"
             "Bu işletim sistemlerindeki yeni keşfedilen güvenlik açıkları (CVE) sonsuza kadar yamasız kalmaktadır. "
             "Ortamda bu sistemlerin varlığı ciddi bir uyumluluk ve siber güvenlik riski teşkil etmektedir.")
    c.fill=_fill("FFCCCC"); c.font=_font(False,"990000",10)
    c.alignment=_align(h="left",wrap=True)
    for r in [23,24,25]: ws5.row_dimensions[r].height=20
    T5=27
    ws5.cell(row=T5-1,column=2,
             value=f"DESTEKLENMEYEN OS KULLANAN CİHAZLAR ({len(eol_df)} adet) – Upgrade Öncelik Listesi").font=_font(True,CD,11)
    if len(eol_df)>0:
        _table(ws5,T5,2,eol_df.sort_values("Risk Skoru",ascending=False),{
            "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı",
            "Sistem":"İşletim Sistemi\n(Destek Sonu Tarihi Geçmiş!)",
            "Offline Gün":"Offline Gün","Risk Skoru":"Risk Skoru","Seviye":"Risk Seviyesi"
        },"784212","Seviye")
    else:
        ws5.cell(row=T5,column=2,value="EoL OS kullanan cihaz tespit edilmedi.").font=_font(False,CG,11)
    for ci,w in enumerate([4,18,15,22,13,12,14],1): ws5.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 5 OK")

    # ── SHEET 6 – Patch & Offline ─────────────────────────────
    ws6=wb.create_sheet("06 Patch ve Offline")
    ws6.sheet_view.showGridLines=False
    _title(ws6,"B2:L3","PATCH & OFFLİNE YÖNETİMİ  |  Hangi Cihaz Ne Kadar Yamasız / Kayıp?","145A32",CW,13)
    patch_df2=df[df["Yamasız Gün"]>60].copy().sort_values("Yamasız Gün",ascending=False)
    offline_df2=df[df["Offline Gün"]>60].copy().sort_values("Offline Gün",ascending=False)
    _kpi(ws6,5,[
        (2,"Yamasız >60g",patch_crit,"784212"),
        (4,"Yamasız >90g",int((df["Yamasız Gün"]>90).sum()),"922B21"),
        (6,"Offline >60g",off_crit,"117A65"),
        (8,"Offline >90g",int((df["Offline Gün"]>90).sum()),"1E8449"),
    ])
    _img(ws6,hist_patch_path,"B8",420,282)
    _img(ws6,hist_off_path,"H8",420,282)

    patch_hdr2={
        "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı","IPAddress":"IP Adresi",
        "Sistem":"İşletim Sistemi","Yamasız Gün":"Yamasız Gün\n(60-90=Kritik | 90+=Acil!)",
        "Offline Gün":"Offline Gün","Risk Skoru":"Risk Skoru","Seviye":"Risk Seviyesi"
    }
    off_hdr2={
        "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı","IPAddress":"IP Adresi",
        "Sistem":"İşletim Sistemi","Offline Gün":"Offline Gün\n(60-90=Kayıp | 90+=Zombi!)",
        "Yamasız Gün":"Yamasız Gün","Risk Skoru":"Risk Skoru","Seviye":"Risk Seviyesi"
    }
    p_cols=["AssetName","Kullanıcı","IPAddress","Sistem","Yamasız Gün","Offline Gün","Risk Skoru","Seviye"]
    T6a=24
    ws6.cell(row=T6a-1,column=2,
             value=f"60+ GÜN YAMASIZ CİHAZLAR ({len(patch_df2)} adet) – Fidye Yazılımı Riski!").font=_font(True,CD,11)
    if len(patch_df2)>0:
        _table(ws6,T6a,2,patch_df2[[c for c in p_cols if c in patch_df2.columns]],
               patch_hdr2,"145A32","Seviye",
               cond_cols={"Yamasız Gün":[(90,9999,"FFCCCC","990000"),(60,90,"FFF3CD","7D3C00")]})
    else:
        ws6.cell(row=T6a,column=2,value="60+ gün yamasız cihaz tespit edilmedi.").font=_font(False,CG,11)
    T6b=T6a+max(len(patch_df2),1)+4
    ws6.cell(row=T6b-1,column=2,
             value=f"60+ GÜN OFFLİNE CİHAZLAR – ZOMBİ RİSKİ ({len(offline_df2)} adet)").font=_font(True,CD,11)
    o_cols=["AssetName","Kullanıcı","IPAddress","Sistem","Offline Gün","Yamasız Gün","Risk Skoru","Seviye"]
    if len(offline_df2)>0:
        _table(ws6,T6b,2,offline_df2[[c for c in o_cols if c in offline_df2.columns]],
               off_hdr2,"117A65","Seviye",
               cond_cols={"Offline Gün":[(90,9999,"FFCCCC","990000"),(60,90,"FFF3CD","7D3C00")]})
    else:
        ws6.cell(row=T6b,column=2,value="60+ gün offline cihaz tespit edilmedi.").font=_font(False,CG,11)
    for ci,w in enumerate([4,16,14,14,18,14,14,12,13],1): ws6.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 6 OK")

    # ── SHEET 7 – Admin & Shadow IT ──────────────────────────
    ws7=wb.create_sheet("07 Admin ve Shadow IT")
    ws7.sheet_view.showGridLines=False
    _title(ws7,"B2:L3","YETKİSİZ ADMİN & SHADOW IT  |  İçeriden Tehdit ve Veri Sızıntısı Riski","922B21",CW,13)
    admin_df=df[df["_RawAdminCount"]>0][["AssetName","Kullanıcı","Kural Dışı Adminler (İsim ve Ünvan)","Risk Skoru","Seviye"]].copy() \
             if "Kural Dışı Adminler (İsim ve Ünvan)" in df.columns else pd.DataFrame()
    shadow_df=df[df["Tespit Edilen Şüpheli Yazılımlar"].notna()][["AssetName","Kullanıcı","Tespit Edilen Şüpheli Yazılımlar","Risk Skoru","Seviye"]].copy() \
              if "Tespit Edilen Şüpheli Yazılımlar" in df.columns else pd.DataFrame()
    _kpi(ws7,5,[
        (2,"Yetkisiz Admin Cihazı",admin_crit,"922B21"),
        (4,"Shadow IT Cihazı",shadow_crit,"1A5276"),
    ])
    # Açıklama
    ws7.merge_cells("F5:K5")
    c=ws7["F5"]; c.value="NEDİR? Yetkisiz Admin = IT onayı olmayan kişi sistem yönetici grubunda → Veri silme/değiştirme/çalma riski!"
    c.fill=_fill("FFCCCC"); c.font=_font(False,"990000",10); c.alignment=_align(h="left",wrap=True)
    ws7.merge_cells("F6:K6")
    c=ws7["F6"]; c.value="NEDİR? Shadow IT = Yetkisiz uzaktan erişim/P2P/hacking araçları → Saldırgan kapısı, veri kaçağı riski!"
    c.fill=_fill("FFF3CD"); c.font=_font(False,"7D3C00",10); c.alignment=_align(h="left",wrap=True)
    ws7.row_dimensions[5].height=22; ws7.row_dimensions[6].height=22

    T7a=8
    ws7.cell(row=T7a-1,column=2,
             value=f"YETKİSİZ ADMİN CİHAZLARI ({len(admin_df)} adet) – Acil İnceleme!").font=_font(True,CD,11)
    if len(admin_df)>0:
        _table(ws7,T7a,2,admin_df.sort_values("Risk Skoru",ascending=False),{
            "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı",
            "Kural Dışı Adminler (İsim ve Ünvan)":"Yetkisiz Admin Listesi\n(IT Onayı Olmayan Yönetici Hesapları)",
            "Risk Skoru":"Risk Skoru","Seviye":"Risk Seviyesi"
        },"922B21","Seviye")
    T7b=T7a+max(len(admin_df),1)+5
    ws7.cell(row=T7b-1,column=2,
             value=f"SHADOW IT / YASAK YAZILIM CİHAZLARI ({len(shadow_df)} adet)").font=_font(True,CD,11)
    if len(shadow_df)>0:
        _table(ws7,T7b,2,shadow_df.sort_values("Risk Skoru",ascending=False),{
            "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı",
            "Tespit Edilen Şüpheli Yazılımlar":"Tespit Edilen Yazılımlar\n(TeamViewer/AnyDesk/Torrent/VPN/Oyun vs.)",
            "Risk Skoru":"Risk Skoru","Seviye":"Risk Seviyesi"
        },"1A5276","Seviye")
    for ci,w in enumerate([4,16,15,38,12,14],1): ws7.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 7 OK")

    # ── SHEET 8 – Yönetici Dashboard ─────────────────────────
    ws8=wb.create_sheet("08 Yonetici Dashboard")
    ws8.sheet_view.showGridLines=False
    _title(ws8,"B2:R3",
           f"YÖNETİCİ ÖZET DASHBOARD  |  {TARIH_H}  |  IT Siber Güvenlik Haftalık Özet",
           "12181F",CW,14)
    ws8.row_dimensions[2].height=32; ws8.row_dimensions[3].height=32
    _img(ws8,pie_path,"B5",380,290)
    _img(ws8,dash_path,"J5",725,290)
    _img(ws8,bar_threat_path,"B24",492,305)
    _img(ws8,bar_user_path,"J24",555,305)

    acil_r=42
    ws8.merge_cells(f"B{acil_r}:R{acil_r}")
    c=ws8[f"B{acil_r}"]
    c.value="ACİL EYLEM PLANI  |  Haftalık Pazartesi Kontrol Listesi"
    c.fill=_fill("922B21"); c.font=_font(True,CW,13); c.alignment=_align()
    ws8.row_dimensions[acil_r].height=28

    hdr_r=acil_r+1
    for col,txt in [(2,"Öncelik"),(3,"Eylem Açıklaması"),(14,"Süre"),(15,"Sorumlu"),(16,"Durum")]:
        c=ws8.cell(row=hdr_r,column=col,value=txt)
        c.fill=_fill("2C3E50"); c.font=_font(True,CW,10); c.alignment=_align(); c.border=_border()
    ws8.row_dimensions[hdr_r].height=24

    eylemler=[
        (43,"KRİTİK",f"Yetkisiz {admin_crit} admin hesabı → BUGÜN AD'den kaldırılmalı, log alınmalı, yöneticiye bildirilmeli.","Bugün","IT Güvenlik","Bekliyor","FFCCCC","990000"),
        (44,"KRİTİK","SEP/Antivirüs eksik cihazlara SCCM/GPO ile paket dağıtımı yapılmalı, tamamlanma raporu alınmalı.","Bu Hafta","IT Altyapı","Bekliyor","FFCCCC","990000"),
        (45,"KRİTİK",f"EoL OS ({eol_count} cihaz: Win7/2008/8.1) → Upgrade takvimi oluştur, yöneticiye sun.","2 Hafta","IT Yöneticisi","Bekliyor","FFCCCC","990000"),
        (46,"YÜKSEK",f"{patch_crit} yamasız cihaza acil Windows Update uygulanmalı. WSUS sağlığı kontrol edilmeli.","Bu Hafta","IT Altyapı","Bekliyor","FFF3CD","7D3C00"),
        (47,"YÜKSEK",f"{shadow_crit} cihazda yasaklı yazılım (TeamViewer/AnyDesk/Torrent) → GPO ile uzaktan kaldırılmalı.","Bu Hafta","IT Güvenlik","Bekliyor","FFF3CD","7D3C00"),
        (48,"YÜKSEK",f"{share_crit} riskli klasör paylaşımı → Herkese açık erişim izni derhal kapatılmalı.","Bu Hafta","IT Altyapı","Bekliyor","FFF3CD","7D3C00"),
        (49,"ORTA",f"{off_crit} offline cihaz → Listele, 90+ gün offline olanları envanterden çıkar.","2 Hafta","IT Destek","Bekliyor","FFFFCC","7D6608"),
        (50,"ORTA","DLP (EDPA) eksik cihazlara kurulum planlanmalı. Deployment paketi hazırlanmalı.","Bu Ay","IT Güvenlik","Bekliyor","FFFFCC","7D6608"),
        (51,"ORTA",f"Kritik disk ({disk_crit} cihaz < %10 boş) → Temizlik script'i çalıştırılmalı.","Bu Hafta","IT Destek","Bekliyor","FFFFCC","7D6608"),
        (52,"DÜŞÜK","WU servisi kapalı cihazlarda GPO ile servis aktif edilmeli, otomatik güncelleme açılmalı.","Bu Ay","IT Altyapı","Bekliyor","D4EDDA","1E8449"),
        (53,"DÜŞÜK","Haftalık rapor direktör ve müdüre sunulmalı. Aksiyonlar takip edilmeli.","Sürekli","IT Yöneticisi","Bekliyor","D4EDDA","1E8449"),
    ]
    for satir,onc,eylem,sure,sorum,durum,bg,fg in eylemler:
        ws8.row_dimensions[satir].height=28
        c1=ws8.cell(row=satir,column=2,value=onc)
        c1.fill=_fill(bg); c1.font=_font(True,fg,9); c1.alignment=_align(); c1.border=_border()
        ws8.merge_cells(start_row=satir,start_column=3,end_row=satir,end_column=13)
        c2=ws8.cell(row=satir,column=3,value=eylem)
        c2.fill=_fill(bg); c2.font=_font(False,fg,10); c2.alignment=_align(h="left",wrap=True); c2.border=_border()
        c3=ws8.cell(row=satir,column=14,value=sure)
        c3.fill=_fill(bg); c3.font=_font(True,fg,9); c3.alignment=_align(); c3.border=_border()
        c4=ws8.cell(row=satir,column=15,value=sorum)
        c4.fill=_fill(bg); c4.font=_font(True,fg,9); c4.alignment=_align(); c4.border=_border()
        c5=ws8.cell(row=satir,column=16,value=durum)
        c5.fill=_fill("F0F0F0"); c5.font=_font(False,"555555",9); c5.alignment=_align(); c5.border=_border()

    note_r=55
    ws8.merge_cells(f"B{note_r}:R{note_r+1}")
    nc=ws8[f"B{note_r}"]
    nc.value=(f"BİLGİ: Bu rapor script çalıştırıldığında otomatik oluşturulmaktadır. "
              f"Veriler Lansweeper'dan SQL sorgusuyla çekilmektedir. "
              f"Alıcılar: {', '.join(MAIL_TO)}  |  CC: {', '.join(MAIL_CC)}")
    nc.fill=_fill("D6EAF8"); nc.font=_font(False,"1A5276",9)
    nc.alignment=_align(h="left",wrap=True); nc.border=_border("2980B9")
    ws8.row_dimensions[note_r].height=16; ws8.row_dimensions[note_r+1].height=16
    for ci,w in enumerate([4,13]+[5]*11+[10,14,12]+[5]*3,1):
        ws8.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 8 OK")

    # ── SHEET 9 – Asset Criticality ───────────────────────────
    ws9=wb.create_sheet("09 Asset Criticality")
    ws9.sheet_view.showGridLines=False
    _title(ws9,"B2:N3",
           "ASSET CRİTİCALİTY ANALİZİ  |  Sunucu/DC/Workstation Risk Önceliklendirmesi",
           "1A5276",CW,13)

    crit_kpi_items=[]
    crit_order=["Domain Controller","Mail Server","Veritabanı Sunucusu","Sunucu","Sunucu (OS)","Laptop","Workstation"]
    crit_colors={"Domain Controller":"7B241C","Mail Server":"922B21","Veritabanı Sunucusu":"C0392B",
                 "Sunucu":"D35400","Sunucu (OS)":"784212","Laptop":"1A5276","Workstation":"2C3E50"}
    col_pos=2
    for ct_lbl in crit_order:
        cnt=int(df["Cihaz_Tipi"].eq(ct_lbl).sum())
        if cnt>0:
            crit_kpi_items.append((col_pos, ct_lbl, cnt, crit_colors.get(ct_lbl,"2C3E50")))
            col_pos+=2
    if crit_kpi_items:
        _kpi(ws9,5,crit_kpi_items)

    ws9.merge_cells("B8:N9")
    c=ws9["B8"]
    c.value=("NASIL ÇALIŞIR?  Lansweeper'dan gelen ham risk skoru, cihazın kurumsal önemine göre çarpanla artırılır. "
             "Domain Controller skoru × 1.6 + 25 bonus. Aynı yamayı kaçırmış bir DC, workstation'dan çok daha riskli! "
             f"Bu hafta {n_promoted} cihazın skoru criticality nedeniyle yukarı çıktı.")
    c.fill=_fill("D6EAF8"); c.font=_font(False,"1A5276",10); c.alignment=_align(h="left",wrap=True)
    ws9.row_dimensions[8].height=22; ws9.row_dimensions[9].height=22

    ws9.merge_cells("B11:N11")
    ws9["B11"].value="ÇARPAN TABLOSU  |  Final Risk = min(100, Lansweeper Skoru × Çarpan + Bonus)"
    ws9["B11"].fill=_fill("1A5276"); ws9["B11"].font=_font(True,CW,11); ws9["B11"].alignment=_align()
    crit_hdr_row=12
    for ci,hdr_t in enumerate(["Cihaz Tipi","Açıklama","Çarpan","Bonus","Skor 50 ise","Final Skor"],2):
        c=ws9.cell(row=crit_hdr_row,column=ci,value=hdr_t)
        c.fill=_fill("2C3E50"); c.font=_font(True,CW,9); c.alignment=_align(); c.border=_border()
    ws9.row_dimensions[crit_hdr_row].height=24
    crit_table_data=[
        ("Domain Controller","Ağın beyni — tüm kimlik doğrulama.",1.6,25,"50×1.6+25=105","→ 100 (Maks!)","7B241C"),
        ("Mail Server","E-posta — hassas veri deposu.",1.5,20,"50×1.5+20=95","→ 95 (Kritik)","922B21"),
        ("Veritabanı Sunucusu","Kurumsal DB — veri ihlali riski.",1.5,20,"50×1.5+20=95","→ 95 (Kritik)","C0392B"),
        ("Sunucu","Genel uygulama/dosya sunucusu.",1.4,15,"50×1.4+15=85","→ 85 (Yüksek)","D35400"),
        ("Sunucu (OS)","Sunucu OS'u çalıştıran cihaz.",1.35,10,"50×1.35+10=77","→ 77 (Yüksek)","784212"),
        ("Laptop","Mobil — dışarı çıkan cihaz.",1.0,0,"50×1.0+0=50","→ 50 (Değişmez)","1A5276"),
        ("Workstation","Standart masaüstü — varsayılan.",1.0,0,"50×1.0+0=50","→ 50 (Değişmez)","2C3E50"),
    ]
    for ri,rd in enumerate(crit_table_data,crit_hdr_row+1):
        bg=rd[6]; ws9.row_dimensions[ri].height=22
        for ci,val in enumerate(rd[:6],2):
            c=ws9.cell(row=ri,column=ci,value=val)
            c.fill=_fill(bg if ci==2 else ("F8F9FA" if ri%2==0 else "FFFFFF"))
            c.font=_font(ci==2,"FFFFFF" if ci==2 else "000000",9)
            c.alignment=_align(h="center"); c.border=_border()

    _img(ws9,crit_chart_path,"B21",650,320)
    _note(ws9,"B38:N39",
          "YORUM: Sunucu/DC sınıfı cihazlar aynı tehdide karşı workstation'lardan çok daha yüksek final skor alır. "
          "IT ekibinin alarm önceliklendirmesi otomatik yapılır — her alarm eşit değildir!","D6EAF8","1A5276")

    T9=42
    crit_tbl_df=df[["AssetName","Kullanıcı","Sistem","Cihaz_Tipi","Crit_Multiplier",
                     "Risk Skoru","Final_Risk_Skoru","Seviye"]].copy()
    crit_tbl_df=crit_tbl_df.sort_values(["Cihaz_Tipi","Final_Risk_Skoru"],ascending=[True,False])
    ws9.cell(row=T9-1,column=2,
             value=f"TÜM CİHAZLAR – HAM SKOR × ÇARPAN = FİNAL SKOR ({n_total} Cihaz)").font=_font(True,CD,11)
    _table(ws9,T9,2,crit_tbl_df,{
        "AssetName":"Cihaz Adı","Kullanıcı":"Kullanıcı","Sistem":"İşletim Sistemi",
        "Cihaz_Tipi":"Cihaz Tipi","Crit_Multiplier":"Çarpan",
        "Risk Skoru":"Ham Skor\n(Lansweeper)","Final_Risk_Skoru":"Final Skor ⭐\n(Criticality)","Seviye":"Seviye"
    },"1A5276","Seviye",
    cond_cols={"Final_Risk_Skoru":[(75,101,"FFCCCC","7B241C"),(50,75,"FFD7D7","990000"),(25,50,"FFF3CD","7D3C00")]})
    for ci,w in enumerate([4,16,14,18,16,8,16,16,13],1): ws9.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 9 (Criticality) OK")

    # ── SHEET 10 – AI Güvenlik Analizi ────────────────────────
    ws10=wb.create_sheet("10 AI Guvenlik Analizi")
    ws10.sheet_view.showGridLines=False
    _title(ws10,"B2:O3",
           f"🤖 AI GÜVENLİK ANALİZİ  |  {TARIH_H}  |  Claude AI Tarafından Otomatik Oluşturuldu",
           "0A0F14",CW,13)

    ws10.merge_cells("B5:C5")
    ws10["B5"].value="🤖 AI"; ws10["B5"].fill=_fill("E74C3C")
    ws10["B5"].font=_font(True,CW,14); ws10["B5"].alignment=_align()
    ws10.merge_cells("D5:O5")
    api_status_lbl = "CANLI AI YORUMU (Claude claude-sonnet-4-20250514)" if CLAUDE_API_KEY else "KURAL TABANLI YORUM (ANTHROPIC_API_KEY eksik)"
    ws10["D5"].value=(f"Kaynak: {api_status_lbl}  |  {TARIH_H}  |  "
                      "Bu sayfa haftalık veriyi analiz eden yapay zeka tarafından otomatik oluşturulur.")
    ws10["D5"].fill=_fill("1C2833"); ws10["D5"].font=_font(False,"AAB7B8",10)
    ws10["D5"].alignment=_align(h="left",wrap=True); ws10.row_dimensions[5].height=28

    ws10.merge_cells("B7:O7")
    ws10["B7"].value="YÖNETİCİ ÖZET  (AI tarafından yazıldı)"
    ws10["B7"].fill=_fill("E74C3C"); ws10["B7"].font=_font(True,CW,12); ws10["B7"].alignment=_align()
    ws10.row_dimensions[7].height=26
    ws10.merge_cells("B8:O13")
    c=ws10["B8"]; c.value=AI_COMMENT_TR
    c.fill=_fill("1C2833"); c.font=_font(False,CW,11)
    c.alignment=_align(h="left",v="top",wrap=True)
    for r in range(8,14): ws10.row_dimensions[r].height=20

    ws10.merge_cells("B15:O15")
    ws10["B15"].value="RİSK TREND DEĞERLENDİRMESİ"
    ws10["B15"].fill=_fill("1A5276"); ws10["B15"].font=_font(True,CW,11); ws10["B15"].alignment=_align()
    ws10.row_dimensions[15].height=24
    ws10.merge_cells("B16:O18")
    c=ws10["B16"]; c.value=AI_RISK_TREND
    c.fill=_fill("1C2833"); c.font=_font(False,"AAB7B8",10)
    c.alignment=_align(h="left",v="top",wrap=True)
    for r in range(16,19): ws10.row_dimensions[r].height=18

    ws10.merge_cells("B20:O20")
    ws10["B20"].value="DETAYLI BULGULAR & ACİL AKSİYONLAR (AI Puanlaması)"
    ws10["B20"].fill=_fill("6C3483"); ws10["B20"].font=_font(True,CW,11); ws10["B20"].alignment=_align()
    ws10.row_dimensions[20].height=24
    for bi,bullet in enumerate(AI_BULLETS_TR,21):
        ws10.merge_cells(f"B{bi}:O{bi}")
        c=ws10[f"B{bi}"]; c.value=bullet
        if bullet.startswith("🔴"):  c.fill=_fill("FFCCCC"); c.font=_font(False,"990000",10)
        elif bullet.startswith("⚡"): c.fill=_fill("FFF3CD"); c.font=_font(False,"7D3C00",10)
        else:                         c.fill=_fill("D4EDDA"); c.font=_font(False,"1E8449",10)
        c.alignment=_align(h="left",wrap=True); ws10.row_dimensions[bi].height=22

    ws10.column_dimensions["B"].width=4
    ws10.column_dimensions["C"].width=110
    for ci in range(4,17): ws10.column_dimensions[get_column_letter(ci)].width=5
    log("  Sheet 10 (AI Analizi) OK")

    # ── SHEET 11 – CVE & Zafiyet İstihbaratı ─────────────────
    ws11 = wb.create_sheet("11 CVE Zafiyet")
    ws11.sheet_view.showGridLines = False
    _title(ws11,"B2:O3",
           f"CVE & ZAFİYET İSTİHBARATI  |  {TARIH_H}  |  NIST NVD Otomatik Tarama",
           "7B241C",CW,13)

    if CVE_DATA:
        # KPI
        _kpi(ws11,5,[
            (2,"Taranan Yazılım",CVE_META.get('toplam_tarama',0),"2C3E50"),
            (4,"Açık Bulunan",cve_vuln_sw,"C0392B"),
            (6,f"Toplam CVE (≥7.0)",cve_total,"922B21"),
            (8,"Kritik (≥9.0)",cve_kritik,"7B241C"),
            (10,"CVE Bonusu Alan Cihaz",int((df['CVE_Bonus']>0).sum()),"6C3483"),
        ])
        tarama_tarihi = CVE_META.get("tarama_tarihi","?")
        ws11.merge_cells("B8:O9")
        c=ws11["B8"]
        c.value=(f"TARAMA TARİHİ: {tarama_tarihi}  |  "
                 "CVE bonusu: Şüpheli yazılım sütunundaki yazılımlar NIST'te arandı, "
                 "bulunan açıklar Final Risk Skoru'na eklendi (Kritik +20, Yüksek +12 puan).  |  "
                 f"Kaynak: {cve_file_used or 'Yok'}")
        c.fill=_fill("FFCCCC"); c.font=_font(False,"7B241C",10)
        c.alignment=_align(h="left",wrap=True,v="top")
        ws11.row_dimensions[8].height=22; ws11.row_dimensions[9].height=22

        # CVE özet tablosu
        ws11.cell(row=11,column=2,
                  value="EN TEHLİKELİ YAZILIMLAR (CVE'ye Göre Sıralı)").font=_font(True,"2C3E50",11)
        hdr_cve=["Yazılım","CVE Sayısı","Max CVSS","Kritiklik","Risk Bonusu","En Tehlikeli CVE ID"]
        for ci,h in enumerate(hdr_cve,2):
            c=ws11.cell(row=12,column=ci,value=h)
            c.fill=_fill("7B241C"); c.font=_font(True,CW,9); c.alignment=_align(); c.border=_border()
        ws11.row_dimensions[12].height=24

        sorted_cve = sorted(CVE_DATA.items(), key=lambda x: x[1].get("max_cvss",0), reverse=True)
        for ri,(sw_name,sw_data) in enumerate(sorted_cve[:25],13):
            ms = float(sw_data.get("max_cvss",0))
            rbg = "FFCCCC" if ms>=9 else "FFF3CD" if ms>=7 else "FFFFFF"
            sev = "KRİTİK" if ms>=9 else "YÜKSEK" if ms>=7 else "-"
            ws11.row_dimensions[ri].height=18
            vals=[sw_name,sw_data.get("cve_sayisi",0),ms,sev,sw_data.get("bonus",0),"→ NIST'te Ara"]
            for ci,val in enumerate(vals,2):
                c=ws11.cell(row=ri,column=ci,value=val)
                c.fill=_fill(rbg); c.font=_font(False,"000000",9)
                c.alignment=_align(h="left" if ci==2 else "center"); c.border=_border()
                if ci==4:  # CVSS renk
                    c.fill=_fill("7B241C" if ms>=9 else "C0392B" if ms>=7 else "27AE60")
                    c.font=_font(True,CW,9)

        # Detaylı Excel varsa bağlantı notu
        cve_excels = sorted(_glob.glob(os.path.join(PROCESSED_DIR,"cve_results_*.xlsx")),reverse=True)
        note_row = 14 + len(sorted_cve[:25])
        ws11.merge_cells(f"B{note_row}:O{note_row}")
        c=ws11[f"B{note_row}"]
        if cve_excels:
            c.value=(f"DETAYLI RAPOR: {cve_excels[0]}  |  "
                     "Bu dosyada her CVE'nin tam açıklaması, vektörü ve CWE bilgisi bulunmaktadır.")
            c.fill=_fill("D6EAF8"); c.font=_font(False,"1A5276",9)
        else:
            c.value="Detaylı CVE raporu için: python scripts/cve_scanner.py"
            c.fill=_fill("FFF3CD"); c.font=_font(False,"7D3C00",9)
        c.alignment=_align(h="left"); ws11.row_dimensions[note_row].height=18

    else:
        # CVE taraması henüz yapılmamış
        ws11.merge_cells("B5:O12")
        c=ws11["B5"]
        c.value=("CVE & Zafiyet taraması henüz çalıştırılmamış.\n\n"
                 "Taramayı başlatmak için:\n\n"
                 "   1. CMD'yi aç\n"
                 "   2. cd C:\\Users\\hsaldiran\\IT_Risk_Engine\\scripts\n"
                 "   3. python cve_scanner.py\n\n"
                 "Tarama süresi: ~5-30 dakika (internet bağlantısına ve yazılım sayısına göre)\n"
                 "Sonraki risk_engine.py çalıştırmasında CVE verileri otomatik dahil edilecek.\n\n"
                 "Opsiyonel: NVD API Key ile çok daha hızlı tarama → https://nvd.nist.gov/developers/request-an-api-key\n"
                 "  set NVD_API_KEY=xxxx-xxxx-xxxx   →   python cve_scanner.py")
        c.fill=_fill("1C2833"); c.font=_font(False,"AAB7B8",11)
        c.alignment=_align(h="left",v="top",wrap=True)
        for r in range(5,13): ws11.row_dimensions[r].height=22

    for ci,w in enumerate([4,28,12,10,12,12,18]+[6]*8,1):
        ws11.column_dimensions[get_column_letter(ci)].width=w
    log("  Sheet 11 (CVE) OK")

    wb.save(OUTPUT_EXCEL)
    log(f"Excel kaydedildi: {OUTPUT_EXCEL}  (11 sheet)")

    # Dashboard icin temiz veri — risk_engine ile birebir ayni sayilar
    # AssetID her zaman dahil: dashboard Lansweeper linklerini buradan üretiyor
    clean_cols = [c for c in [
        "AssetID","AssetName","Kullanıcı","Durum","IPAddress","Sistem",
        "Cihaz_Tipi","Crit_Multiplier",
        "% Boş","Offline Gün","Yamasız Gün",
        "Risk Skoru","Final_Risk_Skoru","CVE_Bonus","Seviye",
        "Risk Analizi","Kural Dışı Adminler (İsim ve Ünvan)",
        "Riskli Paylaşılan Klasörler","Tespit Edilen Şüpheli Yazılımlar",
        "_RawDiskError","_RawAdminCount","_RawUpdateStop",
    ] if c in df.columns]
    # AssetID yoksa boş sütun ekle — dashboard linkler için lazım
    if "AssetID" not in clean_cols:
        df["AssetID"] = ""
        clean_cols.insert(0, "AssetID")
    clean_path = os.path.join(PROCESSED_DIR, "risk_data_current.xlsx")
    df[clean_cols].to_excel(clean_path, index=False)
    log(f"Dashboard verisi kaydedildi: {clean_path}")
    log(f"  → AssetID örnek: {df['AssetID'].dropna().head(3).tolist()}")

    # ═══════════════════════════════════════════════════════════
    # 6. PPTX ALERT (8 slayt — Slide 8: AI + Criticality)
    # ═══════════════════════════════════════════════════════════
    log("PPTX olusturuluyor...")

    # JSON hazırla
    threat_json_list=[json.dumps({"label":str(row["Tehdit"]),"val":int(row["Sayi"])})
                      for _,row in threat_df.head(8).iterrows()]
    threat_json="["+",".join(threat_json_list)+"]"

    top20_list=[]
    for _,r in df[df["Seviye"]=="YUKSEK"].sort_values("Final_Risk_Skoru",ascending=False).head(20).iterrows():
        top20_list.append(json.dumps({
            "name":  str(r.get("AssetName","?"))[:22],
            "user":  str(r.get("Kullanıcı","?"))[:22],
            "ip":    str(r.get("IPAddress","?")),
            "score": int(r.get("Final_Risk_Skoru",0)),
            "ctype": str(r.get("Cihaz_Tipi","Workstation")),
            "risk":  str(r.get("Risk Analizi",""))[:90],
        }))
    top20_json="["+",".join(top20_list)+"]"

    def _jpath(p): return os.path.abspath(p).replace("\\","\\\\")

    pie_abs    = _jpath(pie_path)
    threat_abs = _jpath(bar_threat_path)
    user_abs   = _jpath(bar_user_path)
    dash_abs   = _jpath(dash_path)
    crit_abs   = _jpath(crit_chart_path)
    out_abs    = _jpath(OUTPUT_PPTX)

    # AI yorum — JS string'e gömülecek (tırnakları, emoji ve yeni satırları temizle)
    def _js_str(s):
        import unicodedata
        clean = ""
        for ch in str(s):
            try:
                cat = unicodedata.category(ch)
                # Surrogate veya private-use alanları atla
                if cat.startswith("C") or ord(ch) > 0xFFFF:
                    clean += " "
                else:
                    clean += ch
            except Exception:
                clean += " "
        return (clean
                .replace("\\","\\\\")
                .replace("`","'")
                .replace("${","$[")
                .replace('"','\\"')
                .replace("\n"," ")
                .replace("\r",""))

    ai_comment_js = _js_str(AI_COMMENT_TR[:400] if AI_COMMENT_TR else "AI yorum uretilemedi.")
    # Bullets'tan emoji sembolleri ve unicode dışı karakterleri temizle
    def _clean_bullet(b):
        b2 = _js_str(b)
        for emoji_rep in ["\U0001f534","🔴","\u26a1","⚡","\u2705","✅"]:
            b2 = b2.replace(emoji_rep,"")
        return b2.strip()
    ai_bullets_js = ",".join([f'"{_clean_bullet(b)}"' for b in AI_BULLETS_TR[:6]])

    js_code = f"""const pptxgen=require("pptxgenjs");
const pres=new pptxgen();
pres.layout="LAYOUT_WIDE";
const BG="12181F",RED="E74C3C",OR="F39C12",GR="27AE60",BL="2980B9",WH="FFFFFF",PU="8E44AD";
const nTotal={n_total},nHigh={n_high},nMedium={n_medium},nLow={n_low},tarih="{TARIH_H}";
const pC={patch_crit},oC={off_crit},eol={eol_count},adm={admin_crit},shd={shadow_crit},dsk={disk_crit};
const avg={avg_score};

function kpi(sl,x,y,w,h,lbl,val,bg){{
  sl.addShape(pres.shapes.RECTANGLE,{{x,y,w,h,fill:{{color:bg}},line:{{color:bg}},shadow:{{type:"outer",color:"000000",blur:8,offset:3,angle:135,opacity:0.3}}}});
  sl.addText(val,{{x,y:y+0.03,w,h:h*0.58,fontSize:28,bold:true,color:WH,align:"center",valign:"middle",fontFace:"Arial Black",margin:0}});
  sl.addText(lbl,{{x,y:y+h*0.58,w,h:h*0.42,fontSize:9,color:"BDC3C7",align:"center",fontFace:"Calibri",margin:0}});
}}
function hdr(sl,txt,col){{
  sl.addShape(pres.shapes.RECTANGLE,{{x:0,y:0,w:13.3,h:0.66,fill:{{color:"0A0F14"}},line:{{color:"0A0F14"}}}});
  sl.addText(txt,{{x:0.3,y:0,w:12.5,h:0.66,fontSize:13,bold:true,color:WH,valign:"middle",fontFace:"Arial",margin:0}});
  sl.addShape(pres.shapes.RECTANGLE,{{x:0,y:0.63,w:13.3,h:0.04,fill:{{color:col}},line:{{color:col}}}});
}}

// SLAYT 1: Kapak
let s1=pres.addSlide(); s1.background={{color:BG}};
s1.addShape(pres.shapes.RECTANGLE,{{x:0,y:0,w:0.22,h:7.5,fill:{{color:RED}},line:{{color:RED}}}});
s1.addShape(pres.shapes.RECTANGLE,{{x:0.22,y:0,w:0.06,h:7.5,fill:{{color:"C0392B"}},line:{{color:"C0392B"}}}});
s1.addText("IT SIBER GUVENLIK",{{x:0.45,y:0.75,w:12.5,h:1.0,fontSize:42,bold:true,color:WH,charSpacing:5,fontFace:"Arial Black",margin:0}});
s1.addText("HAFTALIK RISK RAPORU",{{x:0.45,y:1.8,w:12.5,h:0.75,fontSize:28,color:RED,fontFace:"Arial",margin:0}});
s1.addText(tarih+"  |  Lansweeper Otomatik Analizi",{{x:0.45,y:2.65,w:12.5,h:0.45,fontSize:13,color:"7F8C8D",fontFace:"Calibri",margin:0}});
[[0.45,"Toplam Cihaz",String(nTotal),"2C3E50"],[2.9,"Yuksek Risk",String(nHigh),"C0392B"],
 [5.35,"Orta Risk",String(nMedium),"D35400"],[7.8,"Dusuk Risk",String(nLow),"1E8449"],
 [10.25,"EoL OS",String(eol),"6C3483"]].forEach(k=>kpi(s1,k[0],3.3,2.2,1.5,k[1],k[2],k[3]));
[[0.45,"Patch Kritik",String(pC),"784212"],[2.9,"Offline Kritik",String(oC),"117A65"],
 [5.35,"Yetkisiz Admin",String(adm),"922B21"],[7.8,"Shadow IT",String(shd),"1A5276"],
 [10.25,"Ort. Risk Skoru",String(avg),"2C3E50"]].forEach(k=>kpi(s1,k[0],5.0,2.2,1.5,k[1],k[2],k[3]));
s1.addText("Gizlilik: Dahili Kullanim  |  IT Guvenlik Departmani  |  "+tarih,{{x:0.5,y:7.1,w:12,h:0.3,fontSize:9,color:"555555",align:"center",fontFace:"Calibri"}});

// SLAYT 2: Risk Dagilimi
let s2=pres.addSlide(); s2.background={{color:BG}};
hdr(s2,"GENEL RISK DAGILIMI  |  "+String(nTotal)+" Cihaz Analiz Edildi",RED);
s2.addImage({{path:"{pie_abs}",x:0.3,y:0.82,w:5.8,h:4.4}});
[{{y:1.0,lbl:"Yuksek Riskli Cihaz",val:String(nHigh),sub:"Risk Skoru >= 50  |  ACIL MUDAHALE!",bg:"C0392B"}},
 {{y:2.65,lbl:"Orta Riskli Cihaz",val:String(nMedium),sub:"Risk Skoru 25-49  |  Bu Hafta Takip",bg:"D35400"}},
 {{y:4.3,lbl:"Dusuk Riskli Cihaz",val:String(nLow),sub:"Risk Skoru 0-24  |  Rutin Izlem Yeterli",bg:"1E8449"}}].forEach(k=>{{
  s2.addShape(pres.shapes.RECTANGLE,{{x:6.3,y:k.y,w:6.7,h:1.25,fill:{{color:k.bg}},line:{{color:k.bg}},shadow:{{type:"outer",color:"000000",blur:8,offset:3,angle:135,opacity:0.28}}}});
  s2.addText(k.val,{{x:6.3,y:k.y+0.04,w:6.7,h:0.65,fontSize:34,bold:true,color:WH,align:"center",fontFace:"Arial Black",margin:0}});
  s2.addText(k.lbl,{{x:6.3,y:k.y+0.68,w:6.7,h:0.3,fontSize:11,color:"ECF0F1",align:"center",fontFace:"Calibri",margin:0}});
  s2.addText(k.sub,{{x:6.3,y:k.y+0.95,w:6.7,h:0.27,fontSize:9,color:"BDC3C7",align:"center",fontFace:"Calibri",margin:0}});
}});
s2.addText("Toplam "+String(nTotal)+" cihaz analiz edildi.",{{x:6.3,y:5.75,w:6.7,h:0.4,fontSize:11,color:"7F8C8D",align:"center"}});

// SLAYT 3: Tehdit Analizi
let s3=pres.addSlide(); s3.background={{color:BG}};
hdr(s3,"TEHDIT TURU ANALIZI  |  SQL Raporundan Tespit Edilen Riskler",OR);
s3.addImage({{path:"{threat_abs}",x:0.3,y:0.82,w:7.8,h:5.8}});
[{{y:0.85,c:RED,t:"DLP / Yetkisiz Admin",d:"Veri sizintisi ve iceriden tehdit riski. "+String(adm)+" cihazda yetkisiz admin."}},
 {{y:1.9,c:"D35400",t:"Patch Eksikligi",d:String(pC)+" cihaz 60+ gun yamasiz. Kritik CVE aciklari kapanmiyor!"}},
 {{y:2.95,c:PU,t:"Shadow IT / Yasak Yazilim",d:"Uzaktan erisim araclari saldirgan kapisi acabilir. "+String(shd)+" cihazda tespit."}},
 {{y:4.0,c:BL,t:"EoL Isletim Sistemi",d:String(eol)+" cihazda destek omru dolmus OS. Artik yama yok!"}},
 {{y:5.05,c:GR,t:"Zombi / Kayip Cihaz",d:String(oC)+" cihaz 60+ gundur offline. Kim kullaniyor? Nerede?"}},
].forEach(t=>{{
  s3.addShape(pres.shapes.RECTANGLE,{{x:8.3,y:t.y,w:0.07,h:0.82,fill:{{color:t.c}},line:{{color:t.c}}}});
  s3.addText(t.t,{{x:8.45,y:t.y+0.02,w:4.6,h:0.38,fontSize:11,bold:true,color:WH,fontFace:"Arial",margin:0}});
  s3.addText(t.d,{{x:8.45,y:t.y+0.4,w:4.6,h:0.38,fontSize:9,color:"AAB7B8",fontFace:"Calibri",margin:0}});
}});

// SLAYT 4: En Riskli 20 Cihaz
let s4=pres.addSlide(); s4.background={{color:BG}};
hdr(s4,"EN RISKLI 20 CIHAZ  |  Acil Mudahale Listesi",RED);
const devs={top20_json};
const tbl=[[
  {{text:"#",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"Cihaz Adi",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},fontSize:9}}}},
  {{text:"Tip",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"IP Adresi",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"Kullanici",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},fontSize:9}}}},
  {{text:"Final Skor",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"Risk Detayi",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},fontSize:9}}}},
]];
devs.forEach((d,i)=>{{
  const sb=d.score>=75?"7B241C":d.score>=50?"C0392B":"D35400";
  const ctColor=d.ctype&&d.ctype.includes("Controller")?"7B241C":d.ctype&&d.ctype.includes("Sunucu")?"D35400":"1C2833";
  tbl.push([
    {{text:String(i+1),options:{{color:WH,fill:{{color:"1C2833"}},align:"center",fontSize:8}}}},
    {{text:d.name,options:{{color:WH,fill:{{color:"1C2833"}},bold:true,fontSize:8}}}},
    {{text:d.ctype||"WS",options:{{color:WH,fill:{{color:ctColor}},fontSize:7,align:"center"}}}},
    {{text:d.ip,options:{{color:"BDC3C7",fill:{{color:"1C2833"}},fontSize:8}}}},
    {{text:d.user,options:{{color:"BDC3C7",fill:{{color:"1C2833"}},fontSize:8}}}},
    {{text:String(d.score),options:{{color:WH,fill:{{color:sb}},align:"center",bold:true,fontSize:9}}}},
    {{text:d.risk,options:{{color:"BDC3C7",fill:{{color:"1C2833"}},fontSize:7}}}},
  ]);
}});
s4.addTable(tbl,{{x:0.2,y:0.82,w:12.9,h:6.45,border:{{pt:0.5,color:"2C3E50"}},rowH:0.27,colW:[0.3,1.5,1.2,1.4,1.6,0.85,6.05]}});

// SLAYT 5: Kullanici Risk
let s5=pres.addSlide(); s5.background={{color:BG}};
hdr(s5,"KULLANICI DAVRANIŞ RISKI  |  Kimin Cihazlari En Riskli?",PU);
s5.addImage({{path:"{user_abs}",x:0.3,y:0.82,w:8.0,h:5.7}});
s5.addShape(pres.shapes.RECTANGLE,{{x:8.55,y:0.9,w:4.5,h:2.55,fill:{{color:"1C2833"}},line:{{color:"2C3E50"}}}});
s5.addText("PUAN NASIL HESAPLANIYOR?",{{x:8.6,y:0.92,w:4.4,h:0.38,fontSize:10,bold:true,color:WH,fontFace:"Arial",margin:0}});
s5.addText("+ Ort. Cihaz Risk Skoru (0-100)",{{x:8.6,y:1.35,w:4.4,h:0.28,fontSize:9,color:"AAB7B8",fontFace:"Calibri",margin:0}});
s5.addText("+ Yetkisiz Admin Sayisi x 10 puan",{{x:8.6,y:1.62,w:4.4,h:0.28,fontSize:9,color:"AAB7B8",fontFace:"Calibri",margin:0}});
s5.addText("+ 20 puan (herhangi cihaz 60+ gun yamasizsa)",{{x:8.6,y:1.89,w:4.4,h:0.28,fontSize:9,color:"AAB7B8",fontFace:"Calibri",margin:0}});
s5.addText("+ Disk Hata x5  |  WU Kapali x5",{{x:8.6,y:2.16,w:4.4,h:0.28,fontSize:9,color:"AAB7B8",fontFace:"Calibri",margin:0}});
[{{y:3.6,c:"C0392B",l:"50+ puan  →  YUKSEK RISKLI",d:"Derhal incele! Tum cihazlari gozden gecir"}},
 {{y:4.42,c:"D35400",l:"25-49 puan  →  ORTA RISKLI",d:"Bu hafta takip et, eksiklikleri tamamla"}},
 {{y:5.24,c:"1E8449",l:"0-24 puan  →  DUSUK RISKLI",d:"Rutin izlem yeterli"}},
].forEach(r=>{{
  s5.addShape(pres.shapes.RECTANGLE,{{x:8.55,y:r.y,w:4.5,h:0.7,fill:{{color:r.c}},line:{{color:r.c}}}});
  s5.addText(r.l,{{x:8.6,y:r.y+0.04,w:4.4,h:0.32,fontSize:10,bold:true,color:WH,fontFace:"Arial Black",margin:0}});
  s5.addText(r.d,{{x:8.6,y:r.y+0.36,w:4.4,h:0.28,fontSize:8,color:"ECF0F1",fontFace:"Calibri",margin:0}});
}});

// SLAYT 6: Acil Eylem Plani
let s6=pres.addSlide(); s6.background={{color:BG}};
hdr(s6,"ACIL EYLEM PLANI  |  Haftalik Pazartesi Kontrol Listesi",RED);
const acts=[
  {{o:"KRITIK",t:"Yetkisiz "+String(adm)+" admin: AD den kaldir, log al, yoneticiye bildir.",s:"Bugun",r:"IT Guvenlik",bg:"7B241C"}},
  {{o:"KRITIK",t:"EoL OS ("+String(eol)+" cihaz Win7/2008/8.1): upgrade takvimi hazirla.",s:"2 Hafta",r:"IT Yonetici",bg:"7B241C"}},
  {{o:"YUKSEK",t:String(pC)+" yamasiz cihaza acil Win Update. WSUS sagligini kontrol et.",s:"Bu Hafta",r:"IT Altyapi",bg:"D35400"}},
  {{o:"YUKSEK",t:String(shd)+" cihazda yasak yazilim (TeamViewer/Torrent/AnyDesk): GPO ile kaldir.",s:"Bu Hafta",r:"IT Guvenlik",bg:"D35400"}},
  {{o:"YUKSEK",t:"SEP/Antivirus eksik cihazlara SCCM/GPO ile paket dagitimi yap.",s:"Bu Hafta",r:"IT Altyapi",bg:"D35400"}},
  {{o:"ORTA",t:String(oC)+" offline cihazi listele. 90+ gun offline olanlari envanterden cikar.",s:"2 Hafta",r:"IT Destek",bg:"784212"}},
  {{o:"ORTA",t:"DLP (EDPA) eksik cihazlara kurulum planla. Deployment paketini hazirla.",s:"Bu Ay",r:"IT Guvenlik",bg:"784212"}},
  {{o:"DUSUK",t:"Disk kritik ("+String(dsk)+" cihaz <%10 bos): temizlik scripti calistir.",s:"Bu Hafta",r:"IT Destek",bg:"1E8449"}},
];
const atbl=[[
  {{text:"Oncelik",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"Eylem Aciklamasi",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},fontSize:9}}}},
  {{text:"Sure",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
  {{text:"Sorumlu",options:{{bold:true,color:WH,fill:{{color:"1A252F"}},align:"center",fontSize:9}}}},
]];
acts.forEach(a=>atbl.push([
  {{text:a.o,options:{{color:WH,fill:{{color:a.bg}},align:"center",bold:true,fontSize:9}}}},
  {{text:a.t,options:{{color:"ECF0F1",fill:{{color:"1C2833"}},fontSize:9}}}},
  {{text:a.s,options:{{color:WH,fill:{{color:a.bg}},align:"center",fontSize:8}}}},
  {{text:a.r,options:{{color:WH,fill:{{color:"1C2833"}},align:"center",fontSize:8}}}},
]));
s6.addTable(atbl,{{x:0.2,y:0.82,w:12.9,h:6.35,border:{{pt:0.5,color:"2C3E50"}},rowH:0.67,colW:[1.2,8.0,1.4,2.3]}});

// SLAYT 7: Yonetici Ozet
let s7=pres.addSlide(); s7.background={{color:BG}};
hdr(s7,"YONETICI OZET DASHBOARD  |  Tek Bakista Durum",BL);
s7.addImage({{path:"{dash_abs}",x:0.2,y:0.82,w:13,h:4.55}});
[[0.2,"Toplam Cihaz",String(nTotal),"2C3E50"],[2.9,"Yuksek Risk",String(nHigh),"C0392B"],
 [5.6,"Orta Risk",String(nMedium),"D35400"],[8.3,"Dusuk Risk",String(nLow),"1E8449"],
 [11.0,"Ort. Skor",String(avg),"1A5276"]].forEach(k=>kpi(s7,k[0],5.5,2.3,1.75,k[1],k[2],k[3]));

// SLAYT 8: AI Guvenlik Analizi + Criticality
let s8=pres.addSlide(); s8.background={{color:BG}};
hdr(s8,"AI GUVENLIK ANALIZI + CRITICALİTY  |  Zeki Risk Degerlendirmesi","8E44AD");
// Sol: Criticality grafigi
s8.addImage({{path:"{crit_abs}",x:0.2,y:0.82,w:6.8,h:3.5}});
// Sag: AI yorum kutusu
s8.addShape(pres.shapes.RECTANGLE,{{x:7.3,y:0.82,w:5.8,h:3.5,fill:{{color:"1C2833"}},line:{{color:"2C3E50"}}}});
s8.addShape(pres.shapes.RECTANGLE,{{x:7.3,y:0.82,w:5.8,h:0.38,fill:{{color:"8E44AD"}},line:{{color:"8E44AD"}}}});
s8.addText("YAPAY ZEKA YORUMU",{{x:7.35,y:0.83,w:5.7,h:0.35,fontSize:10,bold:true,color:WH,fontFace:"Arial",margin:0}});
s8.addText(`{ai_comment_js}`,{{x:7.35,y:1.25,w:5.7,h:2.0,fontSize:9.5,color:"ECF0F1",fontFace:"Calibri",valign:"top",wrap:true,margin:4}});
// AI bullets
const aiBullets=[{ai_bullets_js}];
const bColors={{"\uD83D\uDD34":"E74C3C","\u26A1":"F39C12","\u2705":"27AE60"}};
aiBullets.slice(0,5).forEach((b,i)=>{{
  const emoji=b.slice(0,2);
  const bc=bColors[emoji]||"555555";
  s8.addShape(pres.shapes.RECTANGLE,{{x:0.2,y:4.45+i*0.52,w:12.9,h:0.46,fill:{{color:"1C2833"}},line:{{color:"2C3E50"}}}});
  s8.addShape(pres.shapes.RECTANGLE,{{x:0.2,y:4.45+i*0.52,w:0.06,h:0.46,fill:{{color:bc}},line:{{color:bc}}}});
  s8.addText(b,{{x:0.32,y:4.47+i*0.52,w:12.6,h:0.42,fontSize:9,color:"ECF0F1",fontFace:"Calibri",valign:"middle",margin:2}});
}});

pres.writeFile({{fileName:"{out_abs}"}}).then(()=>console.log("PPTX OK")).catch(e=>{{console.error("PPTX ERR:",e);process.exit(1);}});
"""
    pptx_js = os.path.join(TEMP_DIR, "gen_pptx.js")
    # Surrogate karakterleri temizle
    js_safe = js_code.encode("utf-8", errors="replace").decode("utf-8", errors="replace")
    with open(pptx_js, "w", encoding="utf-8") as f:
        f.write(js_safe)

    result = subprocess.run(["node", pptx_js], capture_output=True, text=True, timeout=90)
    if result.returncode == 0:
        log(f"PPTX hazır: {OUTPUT_PPTX}")
    else:
        log(f"PPTX HATA: {result.stderr[:600]}")

    # ═══════════════════════════════════════════════════════════
    # 7. OUTLOOK MAIL
    # ═══════════════════════════════════════════════════════════
    if WIN32_OK:
        try:
            log("Outlook mail hazırlanıyor (Display modu)...")
            high_rows = df[df["Seviye"]=="YUKSEK"].sort_values("Risk Skoru",ascending=False).head(10)

            # Kompakt tablo (sadece ilk 10)
            # Yazılım değişim özeti (software_tracker varsa)
            sw_degisim_html = ""
            try:
                import json as _json
                _sw_path = os.path.join(PROCESSED_DIR, "software_changes.json")
                if os.path.exists(_sw_path):
                    _sw = _json.loads(open(_sw_path, encoding="utf-8").read())
                    _n_y = len(_sw.get("yeni_yazilim", []))
                    _n_k = len(_sw.get("kaldirilan", []))
                    _n_g = len(_sw.get("guncellenen", []))
                    _n_s = len(_sw.get("suphe_yeni", []))
                    _sw_tarih = _sw.get("tarih","?")
                    _suphe_rows = ""
                    for _s in _sw.get("suphe_yeni", [])[:5]:
                        _suphe_rows += (f'<tr><td style="border:1px solid #2C3E50;padding:5px 8px;'
                                        f'color:#F85149;background:#1C0808;">{_s["cihaz"]}</td>'
                                        f'<td style="border:1px solid #2C3E50;padding:5px 8px;'
                                        f'color:#FADBD8;background:#1C0808;">{_s["yazilim"]}</td></tr>')
                    _suphe_blok = ""
                    if _n_s > 0:
                        _suphe_blok = f"""<div style="background:#1A0A0A;border-left:4px solid #E74C3C;
padding:10px 16px;margin-top:6px;">
<div style="color:#E74C3C;font-weight:bold;font-size:11px;margin-bottom:6px;">
🚨 {_n_s} ŞÜPHELİ YAZILIM YENİ KURULDU!</div>
<table style="border-collapse:collapse;font-size:11px;width:100%;">
<tr><th style="border:1px solid #2C3E50;padding:4px 8px;color:#BDC3C7;text-align:left;">Cihaz</th>
<th style="border:1px solid #2C3E50;padding:4px 8px;color:#BDC3C7;text-align:left;">Yazılım</th></tr>
{_suphe_rows}</table></div>"""
                    sw_degisim_html = f"""<div style="background:#12181F;padding:14px 22px;margin-top:2px;">
<div style="color:#58A6FF;font-weight:bold;font-size:12px;margin-bottom:8px;">
📦 YAZILIM ENVANTERİ DEĞİŞİMLERİ — {_sw_tarih}</div>
<table style="border-collapse:collapse;font-size:11px;width:auto;">
<tr>
<td style="background:#1E3A1E;color:#3FB950;padding:8px 16px;font-weight:bold;font-size:18px;">{_n_y}</td>
<td style="color:#BDC3C7;padding:8px 12px;">Yeni Kurulan Yazılım</td>
<td style="background:#3A1E1E;color:#E74C3C;padding:8px 16px;font-weight:bold;font-size:18px;">{_n_k}</td>
<td style="color:#BDC3C7;padding:8px 12px;">Kaldırılan Yazılım</td>
<td style="background:#1A1E3A;color:#58A6FF;padding:8px 16px;font-weight:bold;font-size:18px;">{_n_g}</td>
<td style="color:#BDC3C7;padding:8px 12px;">Güncellenen Yazılım</td>
</tr></table>{_suphe_blok}</div>"""
            except Exception:
                sw_degisim_html = ""

            tbl_rows=""
            for _,r in high_rows.iterrows():
                sc=int(r.get("Risk Skoru",0))
                scol="#7B241C" if sc>=75 else "#C0392B"
                an=str(r.get("AssetName","")); aip=str(r.get("IPAddress",""))
                au=str(r.get("Kullanıcı","")); ar=str(r.get("Risk Analizi",""))[:80]
                tbl_rows+=(f'<tr><td style="border:1px solid #2C3E50;padding:6px 8px;color:#ECF0F1;background:#1C2833;font-weight:bold;">{an}</td>'
                           f'<td style="border:1px solid #2C3E50;padding:6px 8px;color:#AAB7B8;background:#1C2833;">{aip}</td>'
                           f'<td style="border:1px solid #2C3E50;padding:6px 8px;color:#AAB7B8;background:#1C2833;">{au}</td>'
                           f'<td style="border:1px solid #2C3E50;padding:6px 8px;text-align:center;background:{scol};color:#FFF;font-weight:bold;">{sc}</td>'
                           f'<td style="border:1px solid #2C3E50;padding:6px 8px;color:#AAB7B8;background:#1C2833;font-size:11px;">{ar}</td></tr>')

            html_body=f"""<html><body style="font-family:Arial,sans-serif;background:#1A252F;margin:0;padding:0;">
<div style="max-width:860px;margin:0 auto;background:#1A252F;">

  <!-- Banner -->
  <div style="background:#12181F;padding:22px 28px;border-left:7px solid #E74C3C;">
    <div style="color:#E74C3C;font-size:11px;letter-spacing:3px;text-transform:uppercase;">IT Güvenlik Departmanı</div>
    <div style="color:#FFFFFF;font-size:22px;font-weight:bold;margin:6px 0 2px 0;letter-spacing:1px;">IT Siber Güvenlik Haftalık Raporu</div>
    <div style="color:#7F8C8D;font-size:12px;">{TARIH_H}  ·  Lansweeper Otomatik Analizi  ·  Dahili Kullanım</div>
  </div>

  <!-- KPI Satırı -->
  <table style="width:100%;border-collapse:collapse;">
    <tr>
      <td style="background:#C0392B;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{n_high}</div><div style="color:#FADBD8;font-size:10px;margin-top:3px;">YÜKSEK RİSKLİ</div></td>
      <td style="background:#D35400;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{n_medium}</div><div style="color:#FAD7A0;font-size:10px;margin-top:3px;">ORTA RİSKLİ</div></td>
      <td style="background:#1E8449;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{n_low}</div><div style="color:#ABEBC6;font-size:10px;margin-top:3px;">DÜŞÜK RİSKLİ</div></td>
      <td style="background:#2C3E50;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{n_total}</div><div style="color:#BDC3C7;font-size:10px;margin-top:3px;">TOPLAM CİHAZ</div></td>
      <td style="background:#6C3483;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{patch_crit}</div><div style="color:#D7BDE2;font-size:10px;margin-top:3px;">PATCH KRİTİK</div></td>
      <td style="background:#784212;padding:16px;text-align:center;"><div style="font-size:30px;font-weight:bold;color:#FFF;">{eol_count}</div><div style="color:#F0B27A;font-size:10px;margin-top:3px;">EOL OS</div></td>
    </tr>
  </table>

  <!-- Pasta grafik -->
  <div style="background:#12181F;padding:18px;text-align:center;">
    <img src="cid:risk_pie" style="width:310px;" alt="Risk Dagilimi"/>
  </div>

  <!-- AI Güvenlik Yorumu -->
  <div style="background:#0D0D1A;border-left:6px solid #8E44AD;padding:16px 22px;margin-top:2px;">
    <div style="color:#8E44AD;font-weight:bold;font-size:13px;margin-bottom:8px;">🤖 AI GÜVENLİK ANALİST YORUMU</div>
    <div style="color:#ECF0F1;font-size:12px;line-height:1.7;">{AI_COMMENT_TR}</div>
  </div>
  <div style="background:#12181F;padding:12px 22px;">
    {''.join([f'<div style="color:{"#C0392B" if b.startswith("🔴") else "#F39C12" if b.startswith("⚡") else "#27AE60"};font-size:11px;margin:4px 0;">{b}</div>' for b in AI_BULLETS_TR[:5]])}
  </div>

  <!-- Acil Uyarı -->
  <div style="background:#1A0A0A;border-left:6px solid #E74C3C;padding:16px 22px;">
    <div style="color:#E74C3C;font-weight:bold;font-size:13px;margin-bottom:8px;">⚠ ACİL MÜDAHALE GEREKTİREN DURUMLAR</div>
    <div style="color:#FADBD8;font-size:12px;margin:5px 0;">• <strong>{admin_crit}</strong> cihazda yetkisiz admin hesabı → <strong>Bu GÜN</strong> AD'den kaldırılmalı!</div>
    <div style="color:#FADBD8;font-size:12px;margin:5px 0;">• <strong>{patch_crit}</strong> cihaz 60+ gün yamasız → Fidye yazılımı saldırısına açık kapı!</div>
    <div style="color:#FADBD8;font-size:12px;margin:5px 0;">• <strong>{eol_count}</strong> cihazda desteksiz OS (Win 7/2008/8.1) → Artık güvenlik yaması çıkmıyor!</div>
    <div style="color:#FADBD8;font-size:12px;margin:5px 0;">• <strong>{shadow_crit}</strong> cihazda yasaklı yazılım (TeamViewer/AnyDesk/Torrent) tespit edildi!</div>
    {'<div style="color:#FADBD8;font-size:12px;margin:5px 0;">• <strong>' + str(cve_kritik) + '</strong> kritik CVE açığı (' + str(cve_vuln_sw) + ' yazılımda) → Acil güncelleme!</div>' if cve_kritik > 0 else ''}
  </div>

  <!-- Yazılım Değişim Özeti -->
  {sw_degisim_html}

  <!-- Bu Hafta Ne Yapılmalı? -->
  <div style="background:#12181F;padding:16px 22px;margin-top:2px;">
    <div style="color:#F39C12;font-weight:bold;font-size:12px;margin-bottom:10px;">📋 BU HAFTA YAPILMASI GEREKENLER</div>
    <table style="width:100%;border-collapse:collapse;font-size:11px;">
      <tr><td style="background:#7B241C;color:#FFF;padding:6px 10px;font-weight:bold;width:90px;">KRİTİK</td>
          <td style="background:#1C2833;color:#ECF0F1;padding:6px 10px;">Yetkisiz admin hesapları: AD'den kaldır, log al → <strong>BUGÜN</strong></td></tr>
      <tr><td style="background:#7B241C;color:#FFF;padding:6px 10px;font-weight:bold;">KRİTİK</td>
          <td style="background:#1A252F;color:#ECF0F1;padding:6px 10px;">EoL OS cihazları için upgrade takvimi hazırla → <strong>2 Hafta</strong></td></tr>
      <tr><td style="background:#D35400;color:#FFF;padding:6px 10px;font-weight:bold;">YÜKSEK</td>
          <td style="background:#1C2833;color:#ECF0F1;padding:6px 10px;">{patch_crit} yamasız cihaza acil Windows Update → <strong>Bu Hafta</strong></td></tr>
      <tr><td style="background:#D35400;color:#FFF;padding:6px 10px;font-weight:bold;">YÜKSEK</td>
          <td style="background:#1A252F;color:#ECF0F1;padding:6px 10px;">Yasaklı yazılımları GPO ile uzaktan kaldır → <strong>Bu Hafta</strong></td></tr>
      <tr><td style="background:#784212;color:#FFF;padding:6px 10px;font-weight:bold;">ORTA</td>
          <td style="background:#1C2833;color:#ECF0F1;padding:6px 10px;">{off_crit} offline cihazı listele ve envanterden çıkar → <strong>2 Hafta</strong></td></tr>
    </table>
  </div>

  <!-- İlk 10 Yüksek Riskli -->
  <div style="padding:0;">
    <div style="background:#12181F;padding:10px 18px;"><span style="color:#E74C3C;font-weight:bold;font-size:12px;">YÜKSEK RİSKLİ CİHAZLAR – İLK 10  (Detaylar Excel raporunda)</span></div>
    <table style="border-collapse:collapse;width:100%;font-size:11px;">
      <tr style="background:#0A0F14;">
        <th style="border:1px solid #2C3E50;padding:7px;color:#BDC3C7;text-align:left;">Cihaz</th>
        <th style="border:1px solid #2C3E50;padding:7px;color:#BDC3C7;">IP</th>
        <th style="border:1px solid #2C3E50;padding:7px;color:#BDC3C7;">Kullanıcı</th>
        <th style="border:1px solid #2C3E50;padding:7px;color:#BDC3C7;text-align:center;">Skor</th>
        <th style="border:1px solid #2C3E50;padding:7px;color:#BDC3C7;">Risk Özeti</th>
      </tr>{tbl_rows}
    </table>
  </div>

  <!-- Alt Bilgi -->
  <div style="background:#0A0F14;padding:12px 22px;border-top:2px solid #E74C3C;margin-top:2px;">
    <p style="color:#555;font-size:10px;margin:0;">Bu e-posta otomatik oluşturulmuştur. Detaylı Excel raporu (8 sheet) ve PPTX sunum ekte sunulmaktadır.<br>
    To: {', '.join(MAIL_TO)}  ·  CC: {', '.join(MAIL_CC)}  ·  {TARIH_H}  ·  IT Güvenlik Departmanı</p>
  </div>

</div></body></html>"""

            pie_abs_m = os.path.abspath(pie_mail_path)
            outlook = win32.Dispatch("outlook.application")
            mail = outlook.CreateItem(0)
            mail.To = "; ".join(MAIL_TO)
            mail.CC = "; ".join(MAIL_CC)
            mail.Subject = (f"[IT GÜVENLİK] Haftalık Risk Raporu | {TARIH_H} | "
                            f"{n_high} Yüksek / {n_medium} Orta / {n_low} Düşük"
                            + (f" {AI_SUBJECT_LINE}" if AI_SUBJECT_LINE else ""))
            mail.HTMLBody = html_body

            att = mail.Attachments.Add(pie_abs_m)
            att.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F","risk_pie")
            mail.Attachments.Add(os.path.abspath(OUTPUT_EXCEL))
            if os.path.exists(OUTPUT_PPTX):
                mail.Attachments.Add(os.path.abspath(OUTPUT_PPTX))

            mail.Display()  # Sadece göster. .Send() ile gönderilir.
            log("Mail hazır. Outlook penceresi açıldı.")
            log("→ Göndermek için Outlook'ta 'Gönder' butonuna basın.")
            log("→ Otomatik göndermek için: mail.Display() → mail.Send()")
        except Exception:
            log(f"OUTLOOK HATASI:\n{traceback.format_exc()}")
            log("Kontrol: Outlook kurulu mu? pywin32 yüklü mü? → pip install pywin32")
    else:
        log("\n[SİMÜLASYON] win32com yok – Windows'ta çalıştırın.")
        log(f"  Konu  : IT GÜVENLİK | {TARIH_H} | {n_high}Y/{n_medium}O/{n_low}D")
        log(f"  To    : {MAIL_TO}")
        log(f"  CC    : {MAIL_CC}")

    # ── ANOMALİ TESPİTİ ─────────────────────────────────────────
    log("Anomali tespiti çalıştırılıyor...")
    try:
        sys.path.insert(0, SCRIPT_DIR)
        from anomaly_engine import anomali_hesapla
        df_with_anomali = anomali_hesapla(df.copy())
        # clean_cols'a anomali sütunlarını ekle ve kaydet
        anomali_cols = ["Anomali_Skoru", "Anomali_Detay", "Anomali_Flag"]
        for col in anomali_cols:
            if col in df_with_anomali.columns:
                df[col] = df_with_anomali[col].values
        # risk_data_current.xlsx'i anomali sütunlarıyla güncelle
        _clean_path2 = os.path.join(PROCESSED_DIR, "risk_data_current.xlsx")
        _all_cols = clean_cols + [c for c in anomali_cols if c in df.columns]
        df[_all_cols].to_excel(_clean_path2, index=False)
        n_anomali = int(df["Anomali_Flag"].sum()) if "Anomali_Flag" in df.columns else 0
        log(f"  Anomali tamamlandı: {n_anomali} anormal cihaz")
    except Exception as _ae:
        log(f"  [UYARI] Anomali atlandı: {_ae}")

    # ── YAZILIM DEĞİŞİM TAKİBİ ─────────────────────────────────
    log("Yazilim degisim takibi calistiriliyor...")
    try:
        from software_tracker import yazilim_guncelle
        _sw_result = yazilim_guncelle()
        if "hata" in _sw_result:
            log(f"  [UYARI] Yazilim takibi: {_sw_result['hata']}")
        else:
            n_s = len(_sw_result.get("suphe_yeni", []))
            log(f"  Yazilim ozet: {_sw_result.get('ozet_str','')}")
            if n_s > 0:
                log(f"  !! {n_s} SUPHE YAZILIM YENİ KURULMUS — dashboard kontrol edin!")
    except FileNotFoundError:
        log("  [BILGI] lansweeper_software.xlsx yok — [2] ile cekin")
    except Exception as _swe:
        log(f"  [UYARI] Yazilim takibi atlandi: {_swe}")

    # ── CİHAZ GEÇMİŞİ KAYDET ───────────────────────────────────
    log("Cihaz gecmisi kaydediliyor...")
    try:
        from device_history_engine import gecmis_kaydet
        _n_hist = gecmis_kaydet(df)
        log(f"  {_n_hist} cihazin gecmisi guncellendi")
    except Exception as _dhe:
        log(f"  [UYARI] Cihaz gecmisi atlandi: {_dhe}")

    # ── HAFTALIK HTML RAPORU ─────────────────────────────────────
    log("HTML raporu üretiliyor...")
    try:
        from report_generator import rapor_olustur
        _hist_file = os.path.join(PROCESSED_DIR, "risk_history.json")
        _hist = json.loads(open(_hist_file, encoding="utf-8").read()) if os.path.exists(_hist_file) else []
        _rapor_path = rapor_olustur(df, cve_data if "cve_data" in dir() else None, _hist)
        if _rapor_path:
            log(f"  Rapor kaydedildi: {_rapor_path}")
    except Exception as _re:
        log(f"  [UYARI] Rapor atlandı: {_re}")

    log("")
    log("="*60)
    log("TAMAMLANDI!")
    log(f"  Excel : {OUTPUT_EXCEL}")
    log(f"  PPTX  : {OUTPUT_PPTX}")
    log(f"  Log   : {LOG_FILE}")
    log("="*60)

except Exception:
    log(f"\n!!! BEKLENMEYEN HATA !!!\n{traceback.format_exc()}")
    log(f"Detaylar için log dosyasına bakın: {LOG_FILE}")
finally:
    # Windows'ta CMD'nin kapanmaması için bekle
    if sys.platform == "win32":
        input("\nDevam etmek için ENTER'a basin...")
