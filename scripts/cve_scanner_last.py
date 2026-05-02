# -*- coding: utf-8 -*-
"""
CVE & Zafiyet İstihbaratı Modülü  —  IT Risk Engine v6.1
=========================================================
KULLANIM:
  1. Lansweeper'dan software_inventory.xlsx çek (aşağıdaki SQL ile)
  2. Bu scripti çalıştır:  python cve_scanner.py
  3. Sonuç:  data/processed/cve_results_YYYYMMDD.xlsx  (risk_engine otomatik okur)

GEREKSİNİMLER:
  pip install pandas openpyxl requests
  (requests yoksa built-in urllib kullanılır)

NIST NVD API:  https://nvd.nist.gov/developers/request-an-api-key
  API key olmadan: 5 istek/30sn
  API key ile:    50 istek/30sn  (ücretsiz, saniyeler içinde alınır)
  NVD_API_KEY ortam değişkeni set edilirse otomatik kullanılır.
"""

import os, sys, json, time, traceback, logging
from datetime import datetime
from pathlib import Path

# ── LOG ─────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
BASE_DIR   = SCRIPT_DIR.parent
LOG_DIR    = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "cve_scanner.log", encoding="utf-8"),
    ]
)
log = logging.getLogger()

import pandas as pd

try:
    import requests
    def _get(url, headers=None, timeout=12):
        r = requests.get(url, headers=headers or {}, timeout=timeout)
        r.raise_for_status()
        return r.json()
    HTTP_LIB = "requests"
except ImportError:
    import urllib.request, urllib.parse
    def _get(url, headers=None, timeout=12):
        req = urllib.request.Request(url, headers=headers or {})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode("utf-8"))
    HTTP_LIB = "urllib"

# ── KONFIGÜRASYON ────────────────────────────────────────────────
INPUT_SW   = BASE_DIR / "data" / "raw"  / "software_inventory.xlsx"   # Lansweeper çıktısı
INPUT_RISK = BASE_DIR / "data" / "raw"  / "lansweeper_risk.xlsx"       # Ana risk dosyası
OUT_DIR    = BASE_DIR / "data" / "processed"
OUT_DIR.mkdir(exist_ok=True)
TARIH      = datetime.now().strftime("%Y%m%d")
OUTPUT_CVE = OUT_DIR / f"cve_results_{TARIH}.xlsx"
CACHE_FILE = OUT_DIR / "cve_cache.json"   # Aynı yazılımı tekrar sorgulamaz

NVD_API_KEY = os.environ.get("NVD_API_KEY", "")
NVD_BASE    = "https://services.nvd.nist.gov/rest/json/cves/2.0"

# API key yoksa: 5 istek/30sn → her istekte 6sn bekle
# API key varsa: 50 istek/30sn → her istekte 0.6sn bekle
RATE_LIMIT_SLEEP = 0.65 if NVD_API_KEY else 6.1

# Öncelikli taranacak yazılımlar (bunlar mutlaka taranır)
PRIORITY_SOFTWARE = [
    "Google Chrome", "Mozilla Firefox", "Microsoft Edge",
    "Java", "Adobe Acrobat", "Adobe Reader", "Adobe Flash",
    "7-Zip", "WinRAR", "WinZip",
    "Microsoft Office", "Microsoft Word", "Microsoft Excel",
    "OpenVPN", "Zoom", "Slack", "Teams",
    "Python", "Node.js", "Git",
    "VLC", "TeamViewer", "AnyDesk",
    "Notepad++", "PuTTY", "WinSCP",
]

# Atlanan yazılımlar (çok fazla sonuç döner, anlamsız)
SKIP_SOFTWARE = {
    "windows", "microsoft corporation", "microsoft windows",
    "intel", "nvidia", "amd", "realtek", "broadcom",
    "vcredist", "visual c++", "directx", "net framework",
    ".net", "redistributable", "runtime", "update",
    "driver", "codec", "dotnet", "windows update",
}

# CVSS eşiği — bu puanın üstündeki CVE'ler raporlanır
CVSS_THRESHOLD = 7.0   # 0-10 arası. 7+ = Yüksek, 9+ = Kritik

# ── CACHE ────────────────────────────────────────────────────────
def load_cache():
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

# ── NVD API SORGUSU ───────────────────────────────────────────────
def query_nvd(software_name: str, cache: dict) -> list[dict]:
    """
    Bir yazılım adı için NVD'den CVE listesi döndürür.
    Sonuçları cache'e kaydeder (aynı yazılımı tekrar sorgulamaz).
    """
    cache_key = software_name.lower().strip()
    if cache_key in cache:
        return cache[cache_key]

    headers = {"User-Agent": "IT-Risk-Engine-CVE-Scanner/1.0"}
    if NVD_API_KEY:
        headers["apiKey"] = NVD_API_KEY

    results = []
    try:
        import urllib.parse
        keyword = urllib.parse.quote(software_name)
        url = f"{NVD_BASE}?keywordSearch={keyword}&resultsPerPage=20"
        data = _get(url, headers=headers, timeout=15)
        time.sleep(RATE_LIMIT_SLEEP)

        for vuln in data.get("vulnerabilities", []):
            cve = vuln.get("cve", {})
            cve_id = cve.get("id", "")
            published = cve.get("published", "")[:10]

            # CVSS skoru al (v3.1 önce, yoksa v2)
            metrics = cve.get("metrics", {})
            cvss_score = None
            severity = ""
            vector = ""
            if "cvssMetricV31" in metrics:
                m = metrics["cvssMetricV31"][0]["cvssData"]
                cvss_score = m.get("baseScore")
                severity   = m.get("baseSeverity", "")
                vector     = m.get("attackVector", "")
            elif "cvssMetricV30" in metrics:
                m = metrics["cvssMetricV30"][0]["cvssData"]
                cvss_score = m.get("baseScore")
                severity   = m.get("baseSeverity", "")
                vector     = m.get("attackVector", "")
            elif "cvssMetricV2" in metrics:
                m = metrics["cvssMetricV2"][0]["cvssData"]
                cvss_score = m.get("baseScore")
                severity   = metrics["cvssMetricV2"][0].get("baseSeverity", "")
                vector     = m.get("accessVector", "")

            if cvss_score is None or float(cvss_score) < CVSS_THRESHOLD:
                continue

            # Açıklama
            desc = ""
            for d in cve.get("descriptions", []):
                if d.get("lang") == "en":
                    desc = d.get("value", "")[:200]
                    break

            # CWE
            cwe = ""
            for w in cve.get("weaknesses", []):
                for wd in w.get("description", []):
                    if wd.get("lang") == "en" and wd.get("value","").startswith("CWE-"):
                        cwe = wd["value"]
                        break

            results.append({
                "CVE_ID":       cve_id,
                "CVSS_Skor":    float(cvss_score),
                "Onem":         severity,
                "Vektor":       vector,  # NETWORK en tehlikelisi
                "Yayin_Tarihi": published,
                "CWE":          cwe,
                "Aciklama":     desc,
            })

    except Exception as e:
        log.warning(f"  NVD sorgu hatası [{software_name}]: {e}")
        time.sleep(RATE_LIMIT_SLEEP * 2)

    # En yüksek skorlu 5'i sakla
    results = sorted(results, key=lambda x: x["CVSS_Skor"], reverse=True)[:5]
    cache[cache_key] = results
    return results

# ── YAZILIM LİSTESİ OLUŞTUR ──────────────────────────────────────
def build_software_list(df_risk: pd.DataFrame) -> list[str]:
    """
    Eğer software_inventory.xlsx varsa onu kullan.
    Yoksa risk dosyasındaki Şüpheli Yazılımlar sütunundan liste çıkar.
    Ayrıca PRIORITY_SOFTWARE listesini mutlaka ekle.
    """
    sw_set = set()

    # 1. software_inventory.xlsx varsa
    if INPUT_SW.exists():
        log.info(f"Yazılım envanteri bulundu: {INPUT_SW}")
        df_sw = pd.read_excel(INPUT_SW)
        # Olası sütun adları
        for col in ["SoftwareName","softwareName","Software","Name","Yazilim","YazilimAdi"]:
            if col in df_sw.columns:
                sw_set.update(df_sw[col].dropna().astype(str).str.strip().tolist())
                log.info(f"  {len(sw_set)} yazılım '{col}' sütunundan yüklendi.")
                break
    else:
        log.warning(f"software_inventory.xlsx bulunamadı: {INPUT_SW}")
        log.warning("Sadece risk dosyasındaki şüpheli yazılımlar + öncelik listesi taranacak.")

    # 2. Şüpheli yazılımlar sütunundan
    col_sw = "Tespit Edilen Şüpheli Yazılımlar"
    if col_sw in df_risk.columns:
        for cell in df_risk[col_sw].dropna():
            for sw in str(cell).split("|"):
                sw_set.add(sw.strip())

    # 3. Öncelik listesini her zaman ekle
    sw_set.update(PRIORITY_SOFTWARE)

    # Filtrele: çok kısa, sayısal, atlanacaklar
    filtered = []
    for sw in sw_set:
        sw_lower = sw.lower().strip()
        if len(sw_lower) < 4:
            continue
        if any(skip in sw_lower for skip in SKIP_SOFTWARE):
            continue
        if sw_lower[0].isdigit():
            continue
        filtered.append(sw.strip())

    # Öncelikli olanlar başta
    priority_set = {p.lower() for p in PRIORITY_SOFTWARE}
    filtered.sort(key=lambda x: (0 if x.lower() in priority_set else 1, x))
    log.info(f"Taranacak benzersiz yazılım sayısı: {len(filtered)}")
    return filtered

# ── CVE SKORU → RİSK BONUSU ──────────────────────────────────────
def cve_risk_bonus(max_cvss: float) -> int:
    """CVE sonucuna göre risk motoruna eklenmek üzere bonus puan."""
    if max_cvss >= 9.0:  return 20  # Kritik
    if max_cvss >= 7.0:  return 12  # Yüksek
    return 5

# ── ANA TARAMA ────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("CVE & Zafiyet İstihbaratı Taraması Başlıyor")
    log.info(f"NVD API Key: {'VAR ✓ (hızlı mod)' if NVD_API_KEY else 'YOK (yavaş mod, ~6sn/sorgu)'}")
    log.info(f"CVSS Eşiği: {CVSS_THRESHOLD}+ (Yüksek & Kritik)")
    log.info("=" * 60)

    if not INPUT_RISK.exists():
        log.error(f"Ana risk dosyası bulunamadı: {INPUT_RISK}")
        sys.exit(1)

    df_risk = pd.read_excel(INPUT_RISK)
    software_list = build_software_list(df_risk)

    cache = load_cache()
    cache_hits = 0
    api_calls  = 0

    # Tarama sonuçları
    vuln_rows    = []   # Her CVE için bir satır
    sw_summary   = []   # Her yazılım için özet

    total = len(software_list)
    for i, sw in enumerate(software_list, 1):
        cache_key = sw.lower().strip()
        from_cache = cache_key in cache

        log.info(f"[{i:3d}/{total}] {sw:<40} {'(cache)' if from_cache else '(API)'}")

        cves = query_nvd(sw, cache)
        if not from_cache:
            api_calls += 1
            save_cache(cache)  # Her çağrıdan sonra kaydet (kesintiye karşı)
        else:
            cache_hits += 1

        if cves:
            max_cvss = max(c["CVSS_Skor"] for c in cves)
            bonus    = cve_risk_bonus(max_cvss)
            sw_summary.append({
                "Yazilim":       sw,
                "CVE_Sayisi":    len(cves),
                "Max_CVSS":      max_cvss,
                "Risk_Bonusu":   bonus,
                "En_Tehlikeli":  cves[0]["CVE_ID"],
                "Aciklama":      cves[0]["Aciklama"][:100],
            })
            for cve in cves:
                vuln_rows.append({
                    "Yazilim":      sw,
                    **cve,
                    "Risk_Bonusu":  bonus,
                })
        else:
            sw_summary.append({
                "Yazilim": sw, "CVE_Sayisi": 0, "Max_CVSS": 0,
                "Risk_Bonusu": 0, "En_Tehlikeli": "", "Aciklama": "",
            })

    # ── CIHAZ-CVE EŞLEŞTİRME ────────────────────────────────────
    # risk dosyasındaki şüpheli yazılım sütunu ile eşleştir
    device_cve_rows = []
    col_sw = "Tespit Edilen Şüpheli Yazılımlar"
    if col_sw in df_risk.columns:
        vuln_sw_set = {r["Yazilim"].lower() for r in vuln_rows}
        for _, row in df_risk.iterrows():
            cell = str(row.get(col_sw, ""))
            if not cell or cell == "nan":
                continue
            device_vulns = []
            for sw_piece in cell.split("|"):
                sw_piece = sw_piece.strip()
                if sw_piece.lower() in vuln_sw_set:
                    matching = [r for r in vuln_rows if r["Yazilim"].lower() == sw_piece.lower()]
                    for m in matching:
                        device_vulns.append({
                            "AssetName": row.get("AssetName", ""),
                            "Kullanici": row.get("Kullanıcı", ""),
                            "Yazilim":   sw_piece,
                            **{k:v for k,v in m.items() if k != "Yazilim"},
                        })
            device_cve_rows.extend(device_vulns)

    # ── ÖZET İSTATİSTİKLER ───────────────────────────────────────
    df_vuln    = pd.DataFrame(vuln_rows)
    df_sw_sum  = pd.DataFrame(sw_summary)
    df_dev_cve = pd.DataFrame(device_cve_rows) if device_cve_rows else pd.DataFrame()

    critical_sw  = df_sw_sum[df_sw_sum["Max_CVSS"] >= 9.0] if len(df_sw_sum) else pd.DataFrame()
    high_sw      = df_sw_sum[(df_sw_sum["Max_CVSS"] >= 7.0) & (df_sw_sum["Max_CVSS"] < 9.0)] if len(df_sw_sum) else pd.DataFrame()
    total_cve    = len(df_vuln)
    vuln_sw_cnt  = int((df_sw_sum["CVE_Sayisi"] > 0).sum()) if len(df_sw_sum) else 0

    log.info("")
    log.info("=" * 60)
    log.info("TARAMA TAMAMLANDI")
    log.info(f"  Taranan yazılım      : {total}")
    log.info(f"  API çağrısı          : {api_calls}  |  Cache hit: {cache_hits}")
    log.info(f"  Açık bulunan yazılım : {vuln_sw_cnt}")
    log.info(f"  Toplam CVE (≥{CVSS_THRESHOLD}) : {total_cve}")
    log.info(f"  Kritik (CVSS ≥9.0)   : {len(critical_sw)}")
    log.info(f"  Yüksek (CVSS 7-9)    : {len(high_sw)}")
    log.info("=" * 60)

    # ── EXCEL ÇIKTISI ─────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _f(h): return PatternFill("solid", fgColor=h)
    def _fn(b=False,c="000000",s=10): return Font(bold=b,color=c,size=s,name="Calibri")
    def _al(h="center",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)
    def _bo(c="CCCCCC"):
        s=Side(border_style="thin",color=c)
        return Border(left=s,right=s,top=s,bottom=s)

    wb = Workbook()
    wb.remove(wb.active)

    # SHEET 1 – Özet Dashboard
    ws1 = wb.create_sheet("01 CVE Ozet")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("B2:L3")
    c=ws1["B2"]
    c.value=f"CVE & ZAFİYET İSTİHBARATI  |  {datetime.now().strftime('%d.%m.%Y')}  |  NIST NVD Otomatik Tarama"
    c.fill=_f("12181F"); c.font=Font(bold=True,color="FFFFFF",size=14,name="Calibri")
    c.alignment=_al(); ws1.row_dimensions[2].height=28; ws1.row_dimensions[3].height=28

    kpis = [
        (2,"Taranan Yazılım",total,"2C3E50"),
        (4,"CVE Bulunan",vuln_sw_cnt,"C0392B"),
        (6,f"Toplam CVE (≥{CVSS_THRESHOLD})",total_cve,"922B21"),
        (8,"Kritik (≥9.0)",len(critical_sw),"7B241C"),
        (10,"Etkilenen Cihaz",len(df_dev_cve),"6C3483"),
    ]
    for col,lbl,val,bg in kpis:
        ws1.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
        ws1.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col+1)
        lc=ws1.cell(row=5,column=col,value=lbl)
        vc=ws1.cell(row=6,column=col,value=val)
        for c2 in [lc,vc]: c2.fill=_f(bg)
        lc.font=_fn(True,"FFFFFF",9); lc.alignment=_al()
        vc.font=_fn(True,"FFFFFF",18); vc.alignment=_al()
    ws1.row_dimensions[5].height=20; ws1.row_dimensions[6].height=42

    ws1.merge_cells("B8:L8")
    c=ws1["B8"]
    c.value=("NASIL OKUNUR?  CVSS 9.0+ = Kritik (kırmızı) → Hemen yama!  |  "
             "CVSS 7.0-8.9 = Yüksek (turuncu) → Bu hafta yama.  |  "
             "Network vektörlü CVE'ler uzaktan istismar edilebilir — en tehlikelisi bunlar!")
    c.fill=_f("D6EAF8"); c.font=_fn(False,"1A5276",10); c.alignment=_al(h="left",w=True)
    ws1.row_dimensions[8].height=24

    # En tehlikeli 20 yazılım tablosu
    ws1.cell(row=10,column=2,value="EN TEHLİKELİ YAZILIMLAR (CVSS'e Göre)").font=_fn(True,"2C3E50",11)
    hdr = ["Yazılım","CVE Sayısı","Max CVSS","Risk Bonusu","En Tehlikeli CVE","Açıklama (ilk 100 karakter)"]
    for ci,h in enumerate(hdr,2):
        c=ws1.cell(row=11,column=ci,value=h)
        c.fill=_f("2C3E50"); c.font=_fn(True,"FFFFFF",9); c.alignment=_al(); c.border=_bo()
    ws1.row_dimensions[11].height=24
    sorted_sw = df_sw_sum[df_sw_sum["CVE_Sayisi"]>0].sort_values("Max_CVSS",ascending=False).head(20)
    for ri,(_, row) in enumerate(sorted_sw.iterrows(), 12):
        ms = float(row["Max_CVSS"])
        rbg = "FFCCCC" if ms>=9 else "FFF3CD" if ms>=7 else "FFFFFF"
        ws1.row_dimensions[ri].height=18
        vals=[row["Yazilim"],int(row["CVE_Sayisi"]),ms,int(row["Risk_Bonusu"]),row["En_Tehlikeli"],row["Aciklama"]]
        for ci,val in enumerate(vals,2):
            cell=ws1.cell(row=ri,column=ci,value=val)
            cell.fill=_f(rbg); cell.font=_fn(False,"000000",9)
            cell.alignment=_al(h="left" if ci>3 else "center",w=False); cell.border=_bo()
    for ci,w in enumerate([4,28,10,10,10,16,50],1):
        ws1.column_dimensions[get_column_letter(ci)].width=w

    # SHEET 2 – Tüm CVE Detayları
    ws2 = wb.create_sheet("02 CVE Detaylari")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("B2:M3")
    ws2["B2"].value="TÜM CVE BULGULARI  |  CVSS ≥7.0 Güvenlik Açıkları"
    ws2["B2"].fill=_f("922B21"); ws2["B2"].font=_fn(True,"FFFFFF",13); ws2["B2"].alignment=_al()
    ws2.row_dimensions[2].height=26; ws2.row_dimensions[3].height=26

    hdr2=["Yazılım","CVE ID","CVSS Skoru","Önem","Vektör (Network=En Tehlikeli)","Yayın Tarihi","CWE","Risk Bonusu","Açıklama"]
    for ci,h in enumerate(hdr2,2):
        c=ws2.cell(row=5,column=ci,value=h)
        c.fill=_f("922B21"); c.font=_fn(True,"FFFFFF",9); c.alignment=_al(w=True); c.border=_bo()
    ws2.row_dimensions[5].height=30

    if len(df_vuln):
        df_vuln_s = df_vuln.sort_values("CVSS_Skor",ascending=False)
        for ri,(_,row) in enumerate(df_vuln_s.iterrows(),6):
            sc=float(row["CVSS_Skor"])
            rbg="FFCCCC" if sc>=9 else "FFF3CD" if sc>=7 else "FFFFFF"
            ws2.row_dimensions[ri].height=16
            vals=[row["Yazilim"],row["CVE_ID"],sc,row["Onem"],row["Vektor"],
                  row["Yayin_Tarihi"],row["CWE"],int(row["Risk_Bonusu"]),row["Aciklama"]]
            for ci,val in enumerate(vals,2):
                c=ws2.cell(row=ri,column=ci,value=val)
                c.fill=_f(rbg); c.font=_fn(False,"000000",9)
                c.alignment=_al(h="center" if ci<=8 else "left"); c.border=_bo()
                if ci==4:  # CVSS rengi
                    c.fill=_f("7B241C" if sc>=9 else "C0392B" if sc>=7 else "27AE60")
                    c.font=_fn(True,"FFFFFF",10)
    for ci,w in enumerate([4,28,16,10,12,24,14,12,12,60],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w

    # SHEET 3 – Cihaz-CVE Eşleşmesi
    ws3 = wb.create_sheet("03 Etkilenen Cihazlar")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("B2:L3")
    ws3["B2"].value="ETKİLENEN CİHAZLAR  |  Yasaklı Yazılım + CVE Açığı = Çift Risk!"
    ws3["B2"].fill=_f("6C3483"); ws3["B2"].font=_fn(True,"FFFFFF",13); ws3["B2"].alignment=_al()
    ws3.row_dimensions[2].height=26; ws3.row_dimensions[3].height=26

    if len(df_dev_cve):
        hdr3=["Cihaz","Kullanıcı","Yazılım","CVE ID","CVSS","Önem","Açıklama"]
        for ci,h in enumerate(hdr3,2):
            c=ws3.cell(row=5,column=ci,value=h)
            c.fill=_f("6C3483"); c.font=_fn(True,"FFFFFF",9); c.alignment=_al(); c.border=_bo()
        ws3.row_dimensions[5].height=24
        df_dc_s=df_dev_cve.sort_values("CVSS_Skor",ascending=False)
        for ri,(_,row) in enumerate(df_dc_s.iterrows(),6):
            sc=float(row["CVSS_Skor"])
            rbg="FFCCCC" if sc>=9 else "FFF3CD" if sc>=7 else "FFFFFF"
            ws3.row_dimensions[ri].height=16
            vals=[row["AssetName"],row["Kullanici"],row["Yazilim"],
                  row["CVE_ID"],sc,row["Onem"],row["Aciklama"]]
            for ci,val in enumerate(vals,2):
                c=ws3.cell(row=ri,column=ci,value=val)
                c.fill=_f(rbg); c.font=_fn(False,"000000",9)
                c.alignment=_al(h="center" if ci>3 else "left"); c.border=_bo()
    else:
        ws3.cell(row=5,column=2,
                 value="Şüpheli yazılım sütunundaki yazılımlar için CVE eşleşmesi bulunamadı veya software_inventory.xlsx yok.").font=_fn(False,"27AE60",11)
    for ci,w in enumerate([4,18,15,22,16,10,12,60],1):
        ws3.column_dimensions[get_column_letter(ci)].width=w

    wb.save(OUTPUT_CVE)
    log.info(f"\nExcel kaydedildi: {OUTPUT_CVE}")
    log.info("risk_engine.py bir sonraki çalıştırmada bu dosyayı otomatik okuyacak.")

    # JSON özet (risk_engine için)
    summary = {
        "tarama_tarihi": TARIH,
        "toplam_tarama": total,
        "vuln_yazilim":  vuln_sw_cnt,
        "toplam_cve":    total_cve,
        "kritik":        len(critical_sw),
        "yuksek":        len(high_sw),
        "sw_risk": {
            row["Yazilim"]: {
                "max_cvss":   row["Max_CVSS"],
                "cve_sayisi": int(row["CVE_Sayisi"]),
                "bonus":      int(row["Risk_Bonusu"]),
            }
            for _,row in df_sw_sum[df_sw_sum["CVE_Sayisi"]>0].iterrows()
        }
    }
    json_path = OUT_DIR / f"cve_summary_{TARIH}.json"
    with open(json_path,"w",encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    log.info(f"JSON özet  : {json_path}")
    log.info("")

if __name__ == "__main__":
    try:
        main()
    except Exception:
        log.error(f"\n!!! HATA !!!\n{traceback.format_exc()}")
    finally:
        if sys.platform == "win32":
            input("\nDevam etmek için ENTER'a basın...")
